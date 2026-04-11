import streamlit as st
import sys
import os
import tempfile
from datetime import datetime as _dt
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent))
import extract_ibkr_data
import calculate_tax_report

st.set_page_config(
    page_title="IBKR Steuerbericht",
    page_icon="🇩🇪",
    layout="centered",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    .stApp {
        background: #0f1117;
        color: #e2e8f0;
    }

    .block-container {
        padding: 1.5rem 1rem 3rem 1rem;
        max-width: 720px;
    }

    /* ── Header ── */
    .page-title {
        font-size: clamp(1.5rem, 5vw, 2rem);
        font-weight: 800;
        color: #f1f5f9;
        letter-spacing: -0.5px;
        margin: 0 0 0.25rem 0;
    }
    .page-sub {
        font-size: 0.9rem;
        color: #64748b;
        margin: 0 0 1.75rem 0;
    }

    /* ── Upload area ── */
    [data-testid="stFileUploader"] {
        background: rgba(255,255,255,0.03);
        border: 1px dashed rgba(255,255,255,0.12);
        border-radius: 14px;
        padding: 0.5rem;
    }

    /* ── Hero card (Zeile 19) ── */
    .hero-card {
        background: linear-gradient(135deg, #1e3a5f 0%, #1a2d4a 100%);
        border: 1px solid rgba(96,165,250,0.25);
        border-radius: 18px;
        padding: 1.75rem 1.5rem;
        margin: 1.5rem 0;
        text-align: center;
    }
    .hero-label {
        font-size: 0.75rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 1.5px;
        color: #60a5fa;
        margin-bottom: 0.5rem;
    }
    .hero-value {
        font-size: clamp(2rem, 8vw, 3rem);
        font-weight: 800;
        letter-spacing: -1px;
        word-break: break-word;
    }
    .hero-formula {
        font-size: 0.8rem;
        color: #64748b;
        margin-top: 0.5rem;
    }

    /* ── Section headers ── */
    .section-title {
        font-size: 0.85rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 2px;
        color: #cbd5e1;
        margin: 2rem 0 0.75rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 1px solid rgba(255,255,255,0.06);
    }

    /* ── Metric grid ── */
    .metric-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
        gap: 0.75rem;
        margin-bottom: 0.75rem;
    }
    .metric-card {
        background: rgba(255,255,255,0.04);
        border: 1px solid rgba(255,255,255,0.08);
        border-radius: 12px;
        padding: 1rem;
        min-width: 0;
    }
    .metric-card.gain  { border-color: rgba(74,222,128,0.2);  background: rgba(74,222,128,0.05); }
    .metric-card.loss  { border-color: rgba(248,113,113,0.2); background: rgba(248,113,113,0.05); }
    .metric-card.saldo { border-color: rgba(250,204,21,0.2);  background: rgba(250,204,21,0.04);  }
    .metric-card.info  { border-color: rgba(148,163,184,0.15); }

    .metric-label {
        font-size: 0.78rem;
        color: #94a3b8;
        margin-bottom: 0.35rem;
        font-weight: 500;
    }
    .metric-value {
        font-size: clamp(1rem, 3.5vw, 1.35rem);
        font-weight: 700;
        word-break: break-word;
        line-height: 1.2;
    }
    .metric-value.green  { color: #4ade80; }
    .metric-value.red    { color: #f87171; }
    .metric-value.yellow { color: #fbbf24; }
    .metric-value.white  { color: #f1f5f9; }

    /* ── Anlage KAP rows ── */
    .kap-row {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 0.85rem 1rem;
        border-radius: 10px;
        margin-bottom: 0.5rem;
        background: rgba(255,255,255,0.03);
        border: 1px solid rgba(255,255,255,0.06);
        gap: 1rem;
        flex-wrap: wrap;
    }
    .kap-left {
        display: flex;
        align-items: center;
        gap: 0.75rem;
        flex: 1;
        min-width: 0;
    }
    .kap-badge {
        font-size: 0.65rem;
        font-weight: 700;
        letter-spacing: 0.5px;
        background: rgba(96,165,250,0.15);
        color: #60a5fa;
        border-radius: 6px;
        padding: 0.2rem 0.45rem;
        white-space: nowrap;
        flex-shrink: 0;
    }
    .kap-desc {
        font-size: 0.9rem;
        color: #cbd5e1;
        font-weight: 500;
        word-break: break-word;
        min-width: 0;
    }
    .kap-value {
        font-size: clamp(0.95rem, 3vw, 1.15rem);
        font-weight: 700;
        text-align: right;
        white-space: nowrap;
        flex-shrink: 0;
    }

    /* highlight row for Zeile 19 */
    .kap-row.highlight {
        background: rgba(96,165,250,0.08);
        border-color: rgba(96,165,250,0.25);
    }
    .kap-row.highlight .kap-badge {
        background: rgba(96,165,250,0.3);
    }
    .kap-row.highlight .kap-desc {
        color: #f1f5f9;
        font-weight: 600;
    }

    /* ── KAP-INV hero card (greenish) ── */
    .hero-card-inv {
        background: linear-gradient(135deg, #1a3a2f 0%, #1a2d2a 100%);
        border: 1px solid rgba(74,222,128,0.25);
        border-radius: 18px;
        padding: 1.75rem 1.5rem;
        margin: 1.5rem 0;
        text-align: center;
    }
    .hero-card-inv .hero-label {
        color: #4ade80;
    }

    /* ── Anlage SO hero card (amber) ── */
    .hero-card-so {
        background: linear-gradient(135deg, #3a2f1a 0%, #2d2a1a 100%);
        border: 1px solid rgba(251,191,36,0.25);
        border-radius: 18px;
        padding: 1.75rem 1.5rem;
        margin: 1.5rem 0;
        text-align: center;
    }
    .hero-card-so .hero-label {
        color: #fbbf24;
    }

    /* ── Audit expander ── */
    [data-testid="stExpander"] {
        border: 1px solid rgba(255,255,255,0.07) !important;
        border-radius: 10px !important;
        background: rgba(255,255,255,0.02) !important;
    }

    /* ── Download button ── */
    [data-testid="stDownloadButton"] > button {
        width: 100%;
        background: rgba(96,165,250,0.1);
        border: 1px solid rgba(96,165,250,0.3);
        color: #60a5fa;
        border-radius: 10px;
        font-weight: 600;
        padding: 0.6rem;
        transition: background 0.2s;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: rgba(96,165,250,0.2);
    }

    /* ── Widget labels (checkbox, selectbox) readable ── */
    [data-testid="stCheckbox"] label p,
    [data-testid="stSelectbox"] label p,
    .stSelectbox label, .stCheckbox label {
        color: #cbd5e1 !important;
    }

    /* hide streamlit branding */
    #MainMenu, footer { visibility: hidden; }
</style>
""", unsafe_allow_html=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def fmt(value: float, decimals: int = 2) -> str:
    s = f"{value:,.{decimals}f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.') + " €"

def fmt_de(value: float, decimals: int = 2) -> str:
    s = f"{value:,.{decimals}f}"
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')

def color_class(value: float) -> str:
    if value > 0:
        return "green"
    elif value < 0:
        return "red"
    return "white"

def metric_card(label: str, value: float, variant: str = "auto") -> str:
    """variant: auto | gain | loss | saldo | info"""
    if variant == "auto":
        variant = "gain" if value > 0 else ("loss" if value < 0 else "info")
    val_color = color_class(value)
    return f"""
    <div class="metric-card {variant}">
        <div class="metric-label">{label}</div>
        <div class="metric-value {val_color}">{fmt(value)}</div>
    </div>"""

def kap_row(zeile: str, label: str, value: float, highlight: bool = False,
            force_positive: bool = False) -> str:
    display_val = abs(value) if force_positive else value
    val_color = "white" if force_positive else color_class(value)
    hl = "highlight" if highlight else ""
    return f"""
    <div class="kap-row {hl}">
        <div class="kap-left">
            <span class="kap-badge">{zeile}</span>
            <span class="kap-desc">{label}</span>
        </div>
        <span class="kap-value" style="color: var(--color-{val_color}, #f1f5f9)">{fmt(display_val)}</span>
    </div>"""

def section_title(text: str):
    st.markdown(f'<div class="section-title">{text}</div>', unsafe_allow_html=True)

def classify_xmls(xml_files):
    """Parse XMLs to extract accountId and date range, group by account.
    Returns dict: accountId -> sorted list of {file, from_date, to_date, name, account_name, currency}
    Latest XML per account = tax year, older = history."""
    import xml.etree.ElementTree as ET
    accounts = {}
    for xml_file in xml_files:
        try:
            content = xml_file.getvalue()
            root = ET.fromstring(content)
            stmt = root.find('.//FlexStatement')
            acct = root.find('.//AccountInformation')
            if stmt is None:
                continue
            account_id = stmt.get('accountId', 'unknown')
            entry = {
                'file': xml_file,
                'from_date': stmt.get('fromDate', ''),
                'to_date': stmt.get('toDate', ''),
                'name': xml_file.name,
                'account_name': acct.get('name', '') if acct is not None else '',
                'currency': acct.get('currency', 'EUR') if acct is not None else 'EUR',
            }
            accounts.setdefault(account_id, []).append(entry)
        except Exception:
            accounts.setdefault('unknown', []).append({
                'file': xml_file, 'from_date': '', 'to_date': '',
                'name': xml_file.name, 'account_name': '', 'currency': 'EUR',
            })
    # Sort each account's XMLs by to_date (latest = tax year)
    for account_id in accounts:
        accounts[account_id].sort(key=lambda x: x['to_date'])
    return accounts

def merge_report_data(reports):
    """Merge multiple report_data dicts (one per account) by summing numeric fields."""
    if not reports:
        return {}
    if len(reports) == 1:
        return reports[0]

    merged = {}

    # Simple sum fields
    for field in ['stocks_gain_eur', 'stocks_loss_eur', 'dividends_eur', 'interest_eur',
                  'debit_interest_eur', 'options_gain_eur', 'options_loss_eur',
                  'withholding_tax_eur', 'fx_total_gain', 'fx_total_loss', 'fx_translation',
                  'fx_correction_total']:
        merged[field] = sum(r.get(field, 0) for r in reports)

    # Recalculated fields
    merged['stocks_net_eur'] = merged['stocks_gain_eur'] + merged['stocks_loss_eur']
    merged['options_net_eur'] = merged['options_gain_eur'] + merged['options_loss_eur']
    merged['topf_1_aktien_netto'] = sum(r.get('topf_1_aktien_netto', 0) for r in reports)
    merged['topf_2_sonstiges_netto'] = sum(r.get('topf_2_sonstiges_netto', 0) for r in reports)
    merged['zeile_19_netto_eur'] = merged['topf_1_aktien_netto'] + merged['topf_2_sonstiges_netto']
    merged['zeile_20_stock_gains_eur'] = sum(r.get('zeile_20_stock_gains_eur', 0) for r in reports)
    merged['zeile_22_other_losses_eur'] = sum(r.get('zeile_22_other_losses_eur', 0) for r in reports)
    merged['zeile_23_stock_losses_eur'] = sum(r.get('zeile_23_stock_losses_eur', 0) for r in reports)
    merged['zeile_41_withholding_tax_eur'] = merged['withholding_tax_eur']

    # Scalars (from first report)
    merged['base_currency'] = reports[0].get('base_currency', 'EUR')
    merged['tax_year'] = reports[0].get('tax_year', 2025)

    # Booleans
    merged['has_trade_price'] = all(r.get('has_trade_price', False) for r in reports)
    merged['xml_has_fx_data'] = all(r.get('xml_has_fx_data', True) for r in reports)
    merged['fx_has_prior_data'] = all(r.get('fx_has_prior_data', False) for r in reports)

    # FX source
    sources = set(r.get('fx_source', 'none') for r in reports)
    merged['fx_source'] = sources.pop() if len(sources) == 1 else 'mixed'

    # fx_results (by currency)
    merged_fx = {}
    for r in reports:
        for curr, data in r.get('fx_results', {}).items():
            if curr not in merged_fx:
                merged_fx[curr] = {'gain': 0, 'loss': 0, 'net': 0}
            for k in ('gain', 'loss', 'net'):
                merged_fx[curr][k] += data.get(k, 0)
    merged['fx_results'] = merged_fx

    # fx_mtm
    merged_mtm = {}
    for r in reports:
        for curr, val in r.get('fx_mtm', {}).items():
            merged_mtm[curr] = merged_mtm.get(curr, 0) + (val or 0)
    merged['fx_mtm'] = merged_mtm

    # Dict sum fields
    for dict_field in ['fx_correction_by_topf', 'fx_corr_gain_adj', 'fx_corr_loss_adj']:
        merged_dict = {}
        for r in reports:
            for k, v in r.get(dict_field, {}).items():
                merged_dict[k] = merged_dict.get(k, 0) + v
        merged[dict_field] = merged_dict

    # FX correction details (list concat)
    merged['fx_correction_details'] = []
    for r in reports:
        merged['fx_correction_details'].extend(r.get('fx_correction_details', []))

    # csv_category_totals (nested dict)
    merged_csv = {}
    for r in reports:
        for cat, data in r.get('csv_category_totals', {}).items():
            if cat not in merged_csv:
                merged_csv[cat] = {}
            for k, v in data.items():
                merged_csv[cat][k] = merged_csv[cat].get(k, 0) + v
    merged['csv_category_totals'] = merged_csv

    # csv_income_totals
    merged_income = {}
    for r in reports:
        for k, v in r.get('csv_income_totals', {}).items():
            merged_income[k] = merged_income.get(k, 0) + v
    merged['csv_income_totals'] = merged_income

    # KAP-INV merge
    merged_kap = {
        'etf_gain_raw_eur': 0, 'etf_loss_raw_eur': 0,
        'etf_gain_taxable_eur': 0, 'etf_loss_taxable_eur': 0,
        'etf_dividends_raw_eur': 0, 'etf_dividends_taxable_eur': 0,
        'etf_wht_eur': 0, 'etf_net_taxable_eur': 0,
        'etf_by_isin': {}, 'etf_unknown_isins': [],
        'etf_stillhalter_premium_eur': 0,
    }
    for r in reports:
        ki = r.get('kap_inv', {})
        for k in ['etf_gain_raw_eur', 'etf_loss_raw_eur', 'etf_gain_taxable_eur',
                   'etf_loss_taxable_eur', 'etf_dividends_raw_eur', 'etf_dividends_taxable_eur',
                   'etf_wht_eur', 'etf_net_taxable_eur', 'etf_stillhalter_premium_eur']:
            merged_kap[k] += ki.get(k, 0)
        for isin, data in ki.get('etf_by_isin', {}).items():
            if isin not in merged_kap['etf_by_isin']:
                merged_kap['etf_by_isin'][isin] = dict(data)
            else:
                existing = merged_kap['etf_by_isin'][isin]
                for nk in ['gain', 'loss', 'div', 'wht', 'gain_taxable', 'loss_taxable', 'div_taxable']:
                    existing[nk] = existing.get(nk, 0) + data.get(nk, 0)
        for isin in ki.get('etf_unknown_isins', []):
            if isin not in merged_kap['etf_unknown_isins']:
                merged_kap['etf_unknown_isins'].append(isin)
    merged['kap_inv'] = merged_kap

    # Anlage SO merge
    merged_so = {
        'total_gain': 0, 'total_loss': 0,
        'taxable_gain': 0, 'taxable_loss': 0,
        'tax_free_gain': 0, 'tax_free_loss': 0,
        'unknown_gain': 0, 'unknown_loss': 0,
        'details': [], 'by_isin': {},
    }
    for r in reports:
        so = r.get('anlage_so', {})
        for k in ['total_gain', 'total_loss', 'taxable_gain', 'taxable_loss',
                   'tax_free_gain', 'tax_free_loss', 'unknown_gain', 'unknown_loss']:
            merged_so[k] += so.get(k, 0)
        merged_so['details'].extend(so.get('details', []))
        for isin, data in so.get('by_isin', {}).items():
            if isin not in merged_so['by_isin']:
                merged_so['by_isin'][isin] = dict(data)
            else:
                existing = merged_so['by_isin'][isin]
                for nk in ['taxable', 'tax_free', 'total']:
                    existing[nk] = existing.get(nk, 0) + data.get(nk, 0)
    merged['anlage_so'] = merged_so

    # Audit merge
    merged_audit = {
        'funds_processed': sum(r.get('audit', {}).get('funds_processed', 0) for r in reports),
        'funds_skipped_year': sum(r.get('audit', {}).get('funds_skipped_year', 0) for r in reports),
        'raw_div_base': sum(r.get('audit', {}).get('raw_div_base', 0) for r in reports),
        'raw_tax_base': sum(r.get('audit', {}).get('raw_tax_base', 0) for r in reports),
        'added_from_summary': sum(r.get('audit', {}).get('added_from_summary', 0) for r in reports),
        'usd_to_eur_rates_count': max(r.get('audit', {}).get('usd_to_eur_rates_count', 0) for r in reports),
        'ecb_rates_used': any(r.get('audit', {}).get('ecb_rates_used', False) for r in reports),
        'stillhalter_count': sum(r.get('audit', {}).get('stillhalter_count', 0) for r in reports),
        'stillhalter_premium_eur': sum(r.get('audit', {}).get('stillhalter_premium_eur', 0) for r in reports),
        'stillhalter_unmatched': [],
        'stillhalter_details': [],
        'cross_year_premium_eur': sum(r.get('audit', {}).get('cross_year_premium_eur', 0) for r in reports),
        'cross_year_by_year': {},
        'cross_year_put_corrections': [],
        'cross_year_put_total': sum(r.get('audit', {}).get('cross_year_put_total', 0) for r in reports),
        'no_invstg_gain': sum(r.get('audit', {}).get('no_invstg_gain', 0) for r in reports),
        'no_invstg_loss': sum(r.get('audit', {}).get('no_invstg_loss', 0) for r in reports),
        'zufluss_premium_eur': sum(r.get('audit', {}).get('zufluss_premium_eur', 0) for r in reports),
        'zufluss_count': sum(r.get('audit', {}).get('zufluss_count', 0) for r in reports),
        'zufluss_details': [],
        'prior_zufluss_correction_eur': sum(r.get('audit', {}).get('prior_zufluss_correction_eur', 0) for r in reports),
        'prior_zufluss_details': [],
        'zufluss_unmatched': [],
    }
    for r in reports:
        a = r.get('audit', {})
        merged_audit['stillhalter_unmatched'].extend(a.get('stillhalter_unmatched', []))
        merged_audit['stillhalter_details'].extend(a.get('stillhalter_details', []))
        merged_audit['cross_year_put_corrections'].extend(a.get('cross_year_put_corrections', []))
        merged_audit['zufluss_details'].extend(a.get('zufluss_details', []))
        merged_audit['prior_zufluss_details'].extend(a.get('prior_zufluss_details', []))
        merged_audit['zufluss_unmatched'].extend(a.get('zufluss_unmatched', []))
        for year, val in a.get('cross_year_by_year', {}).items():
            merged_audit['cross_year_by_year'][year] = merged_audit['cross_year_by_year'].get(year, 0) + val
    merged['audit'] = merged_audit

    # Trade details merge (concatenate lists from all accounts)
    merged['trade_details'] = []
    for r in reports:
        merged['trade_details'].extend(r.get('trade_details', []))
    merged['trade_details'].sort(key=lambda r: r.get('dateTime', '') or r.get('reportDate', '') or 'zzzz')

    return merged

# inline color vars for kap_row values
COLOR_VARS = """
<style>
    :root {
        --color-green: #4ade80;
        --color-red:   #f87171;
        --color-white: #f1f5f9;
    }

    /* Translate Streamlit file uploader to German */
    [data-testid="stFileUploaderDropzoneInstructions"] {
        font-size: 0 !important;
    }
    [data-testid="stFileUploaderDropzoneInstructions"] svg {
        height: 2.5rem;
        width: 2.5rem;
    }
    [data-testid="stFileUploaderDropzoneInstructions"]::after {
        content: "Datei hierher ziehen";
        font-size: 0.875rem;
        color: #e2e8f0;
    }
    [data-testid="stFileUploaderDropzone"] button {
        font-size: 0 !important;
    }
    [data-testid="stFileUploaderDropzone"] button::after {
        content: "Datei auswählen";
        font-size: 0.875rem;
    }
</style>"""


# ── Page ─────────────────────────────────────────────────────────────────────

st.markdown(COLOR_VARS, unsafe_allow_html=True)
st.markdown('<p class="page-title">🇩🇪 IBKR Steuerbericht</p>', unsafe_allow_html=True)
st.markdown('<p class="page-sub">Anlage KAP · Interactive Brokers Flex Query</p>', unsafe_allow_html=True)

st.markdown("""
<div style="background: rgba(251,191,36,0.06); border: 1px solid rgba(251,191,36,0.2); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 0.5rem; font-size: 0.78rem; color: #94a3b8; line-height: 1.6;">
    <strong style="color: #fbbf24;">Wichtiger Hinweis:</strong> Dieses Tool dient ausschließlich als Hilfsmittel zur Aufbereitung von IBKR-Daten für die deutsche Steuererklärung. Es ersetzt keine steuerliche Beratung durch einen Steuerberater oder Wirtschaftsprüfer. Die Ergebnisse sind ohne Gewähr — eine Haftung für die Richtigkeit, Vollständigkeit oder Aktualität der berechneten Werte wird nicht übernommen. Insbesondere bei komplexen Sachverhalten (z.B. Verlustvorträge, Teilfreistellungen, gewerblicher Handel) sollte fachkundiger Rat eingeholt werden.<br><br>
    <strong style="color: #60a5fa;">Datenschutz:</strong> Alle Berechnungen erfolgen ausschließlich lokal. Es werden keine Daten an Server übertragen, gespeichert oder an Dritte weitergegeben.
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div style="background: rgba(59,130,246,0.08); border-left: 3px solid #3b82f6; border-radius: 8px; padding: 0.8rem 1rem; margin-bottom: 0.5rem; font-size: 0.82rem; color: #cbd5e1; line-height: 1.6;">
    <strong style="color: #60a5fa; font-size: 0.9rem;">1. Flex Query XMLs hochladen (Pflicht)</strong><br>
    Alle IBKR Flex Query XMLs auf einmal hochladen: Steuerjahr und ggf. Vorjahre, auch von mehreren Konten.
    Die Konten werden automatisch anhand der Konto-ID erkannt und Vorjahre korrekt zugeordnet.<br>
    <span style="color: #94a3b8; font-size: 0.78rem;">
    <strong>Mehrere Konten:</strong> Jedes Konto wird separat berechnet (eigene Trades, FX, Stillhalter). Die Ergebnisse werden addiert.
    Alle Konten müssen dieselbe Basiswährung (EUR oder USD) haben.<br>
    <strong>Vorjahres-XMLs:</strong> Nur nötig bei Optionen über den Jahreswechsel (Stillhalter-Matching) oder für exakte FX-FIFO-Lots.
    </span><br>
    <span style="color: #64748b;">IBKR &rarr; Performance &amp; Berichte &rarr; Flex-Abfragen &rarr; XML exportieren (gewünschter Zeitraum)</span>
</div>
""", unsafe_allow_html=True)

uploaded_xml_files = st.file_uploader("IBKR Flex Query XMLs hochladen", type="xml",
                                       accept_multiple_files=True, label_visibility="collapsed")

st.markdown("""
<div style="background: rgba(16,185,129,0.08); border-left: 3px solid #10b981; border-radius: 8px; padding: 0.8rem 1rem; margin-bottom: 0.5rem; font-size: 0.82rem; color: #cbd5e1; line-height: 1.6;">
    <strong style="color: #34d399; font-size: 0.9rem;">2. IBKR Standard-Bericht CSV (Plausibilitätscheck)</strong><br>
    <strong style="color: #6ee7b7;">Automatischer Plausibilitätscheck:</strong>
    Der IBKR-Bericht enthält aggregierte Summen pro Kategorie (Aktien, Optionen, Futures, Anleihen, Devisen, Dividenden, Zinsen, Quellensteuer).
    Diese werden automatisch mit unserer Einzelberechnung aus der Flex Query XML verglichen — cent-genaue Übereinstimmung ist das Ziel.<br><br>
    <strong style="color: #6ee7b7;">FX-Fallback:</strong>
    Falls Ihre Flex Query keine <code>FxTransactions</code>-Sektion enthält, liefert der CSV-Bericht die exakten Devisengewinne/-verluste als Ersatz.<br><br>
    <span style="background: rgba(16,185,129,0.12); border-radius: 6px; padding: 0.4rem 0.6rem; display: inline-block; margin-top: 0.2rem; color: #94a3b8;">
    <strong style="color: #a7f3d0;">So erstellen:</strong>
    IBKR &rarr; Performance &amp; Berichte &rarr; Kontoauszüge &rarr;
    <strong>Übersicht: realisierter G&amp;V</strong> &rarr; Zeitraum wählen &rarr; Format: CSV &rarr; Erstellen
    </span><br>
    <span style="color: #94a3b8; font-size: 0.78rem;">
    <strong>Hinweis:</strong> Der CSV-Plausibilitätscheck funktioniert aktuell nur bei einem einzelnen Konto. Bei mehreren Konten wird der CSV-Bericht ignoriert.
    </span>
</div>
""", unsafe_allow_html=True)

ibkr_csv_file = st.file_uploader(
    "IBKR Standard-Bericht (CSV) für Plausibilitätscheck & FX-Fallback",
    type="csv",
    label_visibility="visible")

if not uploaded_xml_files:
    st.stop()

# ── Processing ───────────────────────────────────────────────────────────────

# ── Classify XMLs by account ────────────────────────────────────────────────

accounts = classify_xmls(uploaded_xml_files)
if not accounts:
    st.stop()

# Determine global tax year from latest XML across all accounts
all_to_dates = [xml['to_date'] for xmls in accounts.values() for xml in xmls]
global_tax_year = max(all_to_dates)[:4] if all_to_dates else '2025'

# Split: accounts with a tax-year XML vs. history-only accounts
accounts_to_process = {}
accounts_skipped = []
for acct_id, xmls in accounts.items():
    if xmls[-1]['to_date'][:4] == global_tax_year:
        accounts_to_process[acct_id] = xmls
    else:
        acct_label = xmls[-1]['account_name'] or acct_id
        accounts_skipped.append(f"{acct_label} ({acct_id}, nur bis {xmls[-1]['to_date'][:4]})")

if not accounts_to_process:
    st.error("Keine XML für das Steuerjahr gefunden.")
    st.stop()

if len(accounts_to_process) > 1:
    # Validate: all accounts same base currency
    currencies = {xmls[-1]['currency'] for xmls in accounts_to_process.values()}
    if len(currencies) > 1:
        st.error(f"Unterschiedliche Basiswährungen erkannt: **{', '.join(currencies)}**. "
                 f"Alle Konten müssen dieselbe Basiswährung haben.")
        st.stop()

if len(accounts) > 1:
    acct_info = []
    for acct_id, xmls in sorted(accounts_to_process.items()):
        name = xmls[-1]['account_name'] or acct_id
        years = f"{xmls[0]['from_date'][:4]}–{xmls[-1]['to_date'][:4]}" if len(xmls) > 1 else xmls[0]['to_date'][:4]
        acct_info.append(f"<strong>{name}</strong> ({acct_id}, {len(xmls)} XML{'s' if len(xmls)>1 else ''}, {years})")
    msg = f"<strong style=\"color: #818cf8;\">{len(accounts_to_process)} Konto{'n' if len(accounts_to_process)>1 else ''} für {global_tax_year}</strong>"
    if accounts_skipped:
        msg += f" — {len(accounts_skipped)} übersprungen (kein {global_tax_year}-XML)"
    st.markdown(f"""
<div style="background: rgba(99,102,241,0.08); border: 1px solid rgba(99,102,241,0.25); border-radius: 10px; padding: 0.6rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    {msg}<br>
    {'&ensp;·&ensp;'.join(acct_info)}
</div>
""", unsafe_allow_html=True)

if accounts_skipped:
    st.markdown(f"""
<div style="background: rgba(251,191,36,0.06); border: 1px solid rgba(251,191,36,0.2); border-radius: 10px; padding: 0.6rem 1rem; margin-bottom: 1rem; font-size: 0.78rem; color: #94a3b8;">
    <strong style="color: #fbbf24;">Übersprungen:</strong> {', '.join(accounts_skipped)} — keine Daten für Steuerjahr {global_tax_year} vorhanden.
</div>
""", unsafe_allow_html=True)

# ── Processing ───────────────────────────────────────────────────────────────

with st.spinner("Berechne Steuerreport…"):
    reports = []
    account_names = []

    for acct_id, xmls in sorted(accounts_to_process.items()):
        main_xml = xmls[-1]  # Latest = tax year
        history_xmls = xmls[:-1]  # Older = history
        acct_label = main_xml['account_name'] or acct_id

        with tempfile.TemporaryDirectory() as tmp:
            # Save main XML
            xml_path = os.path.join(tmp, "input.xml")
            with open(xml_path, "wb") as f:
                f.write(main_xml['file'].getbuffer())

            # Save history XMLs
            history_paths = []
            for i, hxml in enumerate(history_xmls):
                hp = os.path.join(tmp, f"history_{i}.xml")
                with open(hp, "wb") as f:
                    f.write(hxml['file'].getbuffer())
                history_paths.append(hp)

            # Save CSV report (only for first account)
            csv_report_path = None
            if ibkr_csv_file is not None and len(reports) == 0:
                csv_report_path = os.path.join(tmp, "ibkr_report.csv")
                with open(csv_report_path, "wb") as f:
                    f.write(ibkr_csv_file.getbuffer())

            try:
                if history_paths:
                    all_xmls_paths = sorted(history_paths) + [xml_path]
                    extract_ibkr_data.extract_fx_multi_xml(all_xmls_paths, tmp)
                else:
                    extract_ibkr_data.parse_ibkr_xml(xml_path, tmp)
                d_acct = calculate_tax_report.calculate_tax(tmp, fx_csv_path=csv_report_path)

                # Validate base currency consistency
                if reports and d_acct.get('base_currency') != reports[0].get('base_currency'):
                    st.error(f"Konto **{acct_label}** hat Basiswährung "
                             f"**{d_acct.get('base_currency')}**, erwartet "
                             f"**{reports[0].get('base_currency')}**.")
                    st.stop()

                reports.append(d_acct)
                account_names.append(acct_label)
            except Exception as e:
                st.error(f"Fehler beim Verarbeiten von Konto **{acct_label}**: {e}")
                st.exception(e)
                st.stop()

    # Merge all accounts
    d = merge_report_data(reports)
    n_accounts = len(reports)

# Derived values
steuerjahr = d.get('tax_year', 2025)
topf_1        = d.get('topf_1_aktien_netto', d.get('stocks_net_eur', 0))
topf_2        = d.get('topf_2_sonstiges_netto',
                      d.get('dividends_eur', 0) + d.get('interest_eur', 0) +
                      d.get('options_gain_eur', 0) + d.get('options_loss_eur', 0))
zeile_19      = d.get('zeile_19_netto_eur', topf_1 + topf_2)
zeile_20      = d.get('zeile_20_stock_gains_eur', d.get('stocks_gain_eur', 0))
zeile_22      = d.get('zeile_22_other_losses_eur', abs(d.get('options_loss_eur', 0)))
zeile_23      = d.get('zeile_23_stock_losses_eur', abs(d.get('stocks_loss_eur', 0)))
quellensteuer = d.get('withholding_tax_eur', 0)

# Zuflussprinzip data
audit = d.get('audit', {})
cross_year_premium = audit.get('cross_year_premium_eur', 0)
cross_year_by_year = audit.get('cross_year_by_year', {})
cross_year_details = [det for det in audit.get('stillhalter_details', []) if det.get('is_cross_year')]
zufluss_details = audit.get('zufluss_details', [])
zufluss_premium = audit.get('zufluss_premium_eur', 0)
prior_zufluss_details = audit.get('prior_zufluss_details', [])
prior_zufluss_correction = audit.get('prior_zufluss_correction_eur', 0)
has_cross_year = len(cross_year_details) > 0 or len(zufluss_details) > 0 or len(prior_zufluss_details) > 0

# ── Flex Query Hinweis ────────────────────────────────────────────────────────

if not d.get('has_trade_price', False):
    st.markdown("""
<div style="background: rgba(251,191,36,0.08); border: 1px solid rgba(251,191,36,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #fbbf24;">Hinweis:</strong> Ihre Flex Query enthält kein <code>tradePrice</code>-Feld. Stillhalterprämien werden mit dem Tagesschlusskurs (<code>closePrice</code>) statt dem tatsächlichen Ausführungspreis berechnet.
    Für genauere Ergebnisse: In IBKR unter <em>Reports → Flex Queries</em> eine erweiterte Query erstellen und bei <em>Trade Confirmation</em> alle Felder aktivieren (insbesondere <em>Execution</em>-Details).
</div>
""", unsafe_allow_html=True)

# ── Stillhalter Warnung ──────────────────────────────────────────────────────

unmatched = d.get('audit', {}).get('stillhalter_unmatched', [])
if unmatched:
    details = ", ".join(f"{u['symbol']} ({u['expiry']})" for u in unmatched)
    st.markdown(f"""
<div style="background: rgba(251,146,60,0.08); border: 1px solid rgba(251,146,60,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #fb923c;">Stillhalter-Warnung:</strong> Für {len(unmatched)} Assignment(s) wurde der ursprüngliche Optionsverkauf nicht gefunden: <strong>{details}</strong>.
    Die Option wurde vermutlich in einem Vorjahr verkauft (Prämie kassiert) und erst im Steuerjahr assigned. Ohne den Original-Trade kann die Prämie nicht berechnet und verbleibt in Topf 1 (Aktien) statt Topf 2 (Sonstiges).
    <br><em>Lösung:</em> Das Vorjahres-XML, in dem die Option verkauft wurde, oben als "Vorjahres-XML" hochladen.
</div>
""", unsafe_allow_html=True)

zufluss_unmatched = d.get('audit', {}).get('zufluss_unmatched', [])
if zufluss_unmatched:
    z_details = ", ".join(f"{u['symbol']}" for u in zufluss_unmatched)
    st.markdown(f"""
<div style="background: rgba(251,146,60,0.08); border: 1px solid rgba(251,146,60,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #fb923c;">Zufluss-Warnung:</strong> {len(zufluss_unmatched)} Glattstellung(en) ohne Eröffnungs-SELL: <strong>{z_details}</strong>.
    Die Option wurde in einem Vorjahr verkauft (Prämie kassiert) und im Steuerjahr geschlossen. Ohne das Vorjahres-XML wird die Prämie doppelt versteuert (einmal im Verkaufsjahr, einmal hier im Schließungsjahr).
    <br><em>Lösung:</em> Das Vorjahres-XML hochladen, damit die Zufluss-Korrektur greifen kann.
</div>
""", unsafe_allow_html=True)

# ── Zuflussprinzip Toggle ────────────────────────────────────────────────────

zuflussprinzip_aktiv = False
if has_cross_year:
    zufluss_parts = []
    if cross_year_details:
        zufluss_parts.append(f"{len(cross_year_details)} Assignment-Prämie(n) aus Vorjahren ({fmt_de(cross_year_premium)} EUR)")
    if zufluss_details:
        zufluss_parts.append(f"{len(zufluss_details)} offene Stillhalter-Position(en) mit Zufluss im Steuerjahr ({fmt_de(zufluss_premium)} EUR, bereits in Berechnung enthalten)")
    if prior_zufluss_details:
        zufluss_parts.append(f"{len(prior_zufluss_details)} Vorjahres-Prämie(n) aus Glattstellungen korrigiert (-{fmt_de(prior_zufluss_correction)} EUR, bereits in Berechnung enthalten)")
    st.markdown(f"""
<div style="background: rgba(168,85,247,0.08); border: 1px solid rgba(168,85,247,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #a855f7;">Zuflussprinzip (BMF Rn. 25, 33):</strong> {"<br>".join(zufluss_parts)}
</div>
""", unsafe_allow_html=True)
    if cross_year_details:
        zuflussprinzip_aktiv = st.checkbox(
            "Zuflussprinzip anwenden (BMF Rn. 25, 33)",
            value=False,
            help="Verschiebt Assignment-Prämien aus Vorjahren aus dem aktuellen Steuerjahr heraus. "
                 "Diese Prämien gehören in die Steuererklärung des jeweiligen Vorjahres. "
                 "Offene Stillhalter-Positionen und Vorjahres-Korrekturen sind bereits automatisch berechnet.")

# Adjusted values for Zuflussprinzip
# Note: zufluss_premium and prior_zufluss_correction are already applied in calculate_tax_report.py
# The toggle only adjusts assignment cross-year premiums (existing behavior)
adj_cross = cross_year_premium if zuflussprinzip_aktiv else 0
adj_topf_2 = topf_2 - adj_cross
adj_zeile_19 = zeile_19 - adj_cross

# ── InvStG Toggle (vor Hero, damit Adjustments in Zeile 19 einfließen) ───────

kap_inv = d.get('kap_inv', {})
etf_by_isin = kap_inv.get('etf_by_isin', {})
has_etf_data = len(etf_by_isin) > 0

invstg_aktiv = False
if has_etf_data:
    n_aktien = sum(1 for v in etf_by_isin.values() if v.get('classification') == 'aktienfonds')
    n_sonst = sum(1 for v in etf_by_isin.values() if v.get('classification') != 'aktienfonds')
    etf_tickers = ", ".join(sorted(v.get('ticker', '?') for v in etf_by_isin.values()))

    st.markdown(f"""
<div style="background: rgba(74,222,128,0.08); border: 1px solid rgba(74,222,128,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #4ade80;">InvStG-Klassifizierung (§2 InvStG):</strong> {len(etf_by_isin)} ETFs in Ihrer XML erkannt, die nach dem Investmentsteuergesetz als Investmentfonds gelten und auf <strong>Anlage KAP-INV</strong> gemeldet werden (nicht auf Anlage KAP).
    Davon {n_aktien} Aktienfonds (30% TFS) und {n_sonst} sonstige Fonds (0% TFS).<br>
    <span style="color: #64748b; font-size: 0.75rem;">Betroffene ETFs: {etf_tickers}</span>
</div>
""", unsafe_allow_html=True)
    invstg_aktiv = st.checkbox(
        "InvStG-Klassifizierung anwenden (Anlage KAP-INV)",
        value=True,
        help="ETFs werden als Investmentfonds nach InvStG behandelt: separate Meldung auf Anlage KAP-INV, "
             "Teilfreistellung (30% für Aktienfonds, 0% für sonstige). "
             "Deaktivieren = alle ETFs wie normale Aktien auf Anlage KAP behandeln.")

    if not invstg_aktiv:
        etf_raw_net = kap_inv.get('etf_gain_raw_eur', 0) + kap_inv.get('etf_loss_raw_eur', 0)
        topf_1 += etf_raw_net
        zeile_19 += etf_raw_net
        adj_zeile_19 += etf_raw_net
        zeile_20 += max(kap_inv.get('etf_gain_raw_eur', 0), 0)
        zeile_23 += abs(min(kap_inv.get('etf_loss_raw_eur', 0), 0))
        adj_topf_2 += kap_inv.get('etf_dividends_raw_eur', 0)
        adj_zeile_19 += kap_inv.get('etf_dividends_raw_eur', 0)
        zeile_19 += kap_inv.get('etf_dividends_raw_eur', 0)
        quellensteuer += kap_inv.get('etf_wht_eur', 0)

# ── Tageskurs-Korrektur (§20 Abs. 4 S. 1 EStG) ──────────────────────────────

fx_corr_total = d.get('fx_correction_total', 0)
fx_corr_by_topf = d.get('fx_correction_by_topf', {})
tk_gain_adj = d.get('fx_corr_gain_adj', {})
tk_loss_adj = d.get('fx_corr_loss_adj', {})
tageskurs_aktiv = False
tageskurs_kapinv_corr = 0

if abs(fx_corr_total) > 0.01:
    st.markdown(f"""
<div style="background: rgba(251,146,60,0.08); border: 1px solid rgba(251,146,60,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8; line-height: 1.6;">
    <strong style="color: #fb923c;">Tageskurs-Methode (§20 Abs. 4 S. 1 EStG):</strong>
    <em>"Bei nicht in Euro getätigten Geschäften sind die Einnahmen im Zeitpunkt der Veräußerung und die Anschaffungskosten im Zeitpunkt der Anschaffung in Euro umzurechnen."</em><br>
    IBKR rechnet den gesamten Netto-Gewinn zum Schlusskurs um. Das Gesetz verlangt: Erlös zum Verkaufskurs, Kosten zum Kaufkurs.
    Abweichung für {steuerjahr}: <strong style="color: #fb923c;">{"+" if fx_corr_total >= 0 else ""}{fmt_de(fx_corr_total)} EUR</strong> (CLOSED_LOT Analyse, ohne Futures*).<br>
    <span style="color: #64748b; font-size: 0.72rem;">*Futures ausgeschlossen: Die Kostenbasis bei Futures ist der volle Kontraktwert (z.B. $200.000 bei ZT), nicht die tatsächlich gezahlte Margin. Eine FX-Korrektur auf den Notional würde massive Phantom-Gewinne/-Verluste erzeugen, die keinem realen Cashflow entsprechen.</span>
</div>
""", unsafe_allow_html=True)
    tageskurs_aktiv = st.checkbox(
        "Tageskurs-Methode anwenden (§20 Abs. 4 S. 1 EStG)",
        value=False,
        help="Rechnet Veräußerungserlöse und Anschaffungskosten jeweils zum FX-Kurs ihres eigenen Datums um, "
             "statt den gesamten Netto-PnL zum Schlusskurs. Gesetzlich korrekt, aber Abweichung zur IBKR-Methode. "
             "Nur verfügbar mit Extended Flex Query (CLOSED_LOT Daten).")

    if tageskurs_aktiv:
        corr_topf1 = fx_corr_by_topf.get('Topf1', 0)
        corr_topf2 = fx_corr_by_topf.get('Topf2', 0)
        tageskurs_kapinv_corr = fx_corr_by_topf.get('KAP-INV', 0)
        topf_1 += corr_topf1
        adj_topf_2 += corr_topf2
        if invstg_aktiv:
            adj_zeile_19 += corr_topf1 + corr_topf2
            zeile_19 += corr_topf1 + corr_topf2
        else:
            adj_zeile_19 += fx_corr_total
            zeile_19 += fx_corr_total
        # Zeilen 20/22/23 per-Lot gain/loss adjustments
        zeile_20 += tk_gain_adj.get('Topf1', 0)
        zeile_23 -= tk_loss_adj.get('Topf1', 0)
        zeile_22 -= tk_loss_adj.get('Topf2', 0)
    else:
        tageskurs_kapinv_corr = 0

# ── Basiswährung ────────────────────────────────────────────────────────────

base_curr = d.get('base_currency', 'USD')
base_icon = "🇪🇺" if base_curr == "EUR" else "🇺🇸"
st.markdown(f"""
<div style="background: rgba(99,102,241,0.08); border: 1px solid rgba(99,102,241,0.25); border-radius: 10px; padding: 0.6rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    {base_icon} <strong style="color: #818cf8;">Basiswährung: {base_curr}</strong> — {"Beträge in StmtFunds sind bereits in EUR (BaseCurrency-Ansicht)." if base_curr == "EUR" else "USD-Beträge werden über tägliche Wechselkurse in EUR umgerechnet."}
</div>
""", unsafe_allow_html=True)

if n_accounts > 1:
    st.markdown(f"""
<div style="background: rgba(251,191,36,0.08); border: 1px solid rgba(251,191,36,0.25); border-radius: 10px; padding: 0.6rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #fbbf24;">{n_accounts} Konten zusammengeführt</strong> — Jedes Konto wurde separat berechnet, die Ergebnisse wurden addiert.
</div>
""", unsafe_allow_html=True)

# ── FxTransactions-Warnung ───────────────────────────────────────────────────

xml_has_fx = d.get('xml_has_fx_data', True)
fx_source = d.get('fx_source', 'none')

if not xml_has_fx and fx_source != 'csv':
    st.markdown("""
<div style="background: rgba(251,146,60,0.1); border: 1px solid rgba(251,146,60,0.35); border-radius: 10px; padding: 0.85rem 1rem; margin-bottom: 1rem; font-size: 0.82rem; color: #cbd5e1; line-height: 1.6;">
    <strong style="color: #fb923c; font-size: 0.9rem;">Flex Query unvollständig: Keine FX-Transaktionsdaten</strong><br>
    Ihre Flex Query XML enthält keine <code>FxTransactions</code>-Sektion. Ohne diese Daten können Fremdwährungs-Gewinne/-Verluste
    nur approximiert werden (FIFO-Schätzung mit eingeschränkter Genauigkeit).<br><br>
    <strong style="color: #fdba74;">Lösung:</strong> Laden Sie den <strong>IBKR Standard-Bericht (CSV)</strong> oben hoch —
    dieser enthält die exakten Devisengewinne/-verluste, die IBKR intern per FIFO berechnet.<br>
    <span style="color: #94a3b8; font-size: 0.78rem;">
    Alternativ: In IBKR unter <em>Reports → Flex Queries → Configure</em> die Sektion
    <em>FX Transactions</em> aktivieren und die XML neu exportieren.
    </span>
</div>
""", unsafe_allow_html=True)
elif not xml_has_fx and fx_source == 'csv':
    st.markdown("""
<div style="background: rgba(16,185,129,0.08); border: 1px solid rgba(16,185,129,0.25); border-radius: 10px; padding: 0.6rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #34d399;">FX-Daten aus CSV übernommen.</strong>
    Die Flex Query enthält keine FxTransactions, aber der IBKR Standard-Bericht liefert exakte Devisenwerte.
</div>
""", unsafe_allow_html=True)

# ── Hero ─────────────────────────────────────────────────────────────────────

hero_color = "#4ade80" if adj_zeile_19 >= 0 else "#f87171"
invstg_label = ""
st.markdown(f"""
<div class="hero-card">
    <div class="hero-label">Zeile 19 · Ausländische Kapitalerträge (Netto){"  · Zuflussprinzip" if zuflussprinzip_aktiv else ""}{invstg_label}</div>
    <div class="hero-value" style="color:{hero_color}">{fmt(adj_zeile_19)}</div>
    <div class="hero-formula">Topf 1 ({fmt(topf_1)}) + Topf 2 ({fmt(adj_topf_2)})</div>
</div>
""", unsafe_allow_html=True)

# ── Topf 1: Aktien ───────────────────────────────────────────────────────────

topf_1_label = "Topf 1 · Aktien ohne ETF-Fonds (separate Verrechnung §20 Abs. 6 S. 4 EStG)" if (has_etf_data and invstg_aktiv) else "Topf 1 · Aktien (separate Verrechnung §20 Abs. 6 S. 4 EStG)"
section_title(topf_1_label)

st.markdown(
    '<div class="metric-grid">'
    + metric_card("Aktiengewinne", d['stocks_gain_eur'] + (tk_gain_adj.get('Topf1', 0) if tageskurs_aktiv else 0), "gain")
    + metric_card("Aktienverluste", d['stocks_loss_eur'] + (tk_loss_adj.get('Topf1', 0) if tageskurs_aktiv else 0), "loss")
    + metric_card("Saldo Aktien", topf_1, "saldo")
    + '</div>',
    unsafe_allow_html=True
)

# ── Cross-Year Put-Korrektur ─────────────────────────────────────────────────

cross_put_corrections = d.get('audit', {}).get('cross_year_put_corrections', [])
cross_put_total = d.get('audit', {}).get('cross_year_put_total', 0)
if cross_put_corrections:
    st.markdown(f"""
<div style="background: rgba(168,85,247,0.08); border: 1px solid rgba(168,85,247,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #a855f7;">Put-Assignment Korrektur (BMF Rn. 33):</strong>
    {len(cross_put_corrections)} Aktienverkäufe stammen aus Put-Assignments früherer Jahre.
    Die Prämie ({fmt_de(cross_put_total)} EUR) wurde bereits im Assignment-Jahr versteuert und wird
    hier vom Aktien-PnL abgezogen (Einstandskurs = Strike, nicht Strike minus Prämie).
</div>
""", unsafe_allow_html=True)
    with st.expander(f"Details: {len(cross_put_corrections)} Cross-Year Put-Korrekturen"):
        put_table = "| Symbol | Shares | Strike | Korrektur | Aus Jahr |\n|--------|--------|--------|-----------|----------|\n"
        for c in cross_put_corrections:
            put_table += f"| {c['symbol']} | {c['shares']} | {c['strike']} | {fmt_de(c['correction_eur'])} EUR | {c['assignment_year']} |\n"
        put_table += f"| **Gesamt** | | | **{fmt_de(cross_put_total)} EUR** | |\n"
        st.markdown(put_table)

# ── Topf 2: Sonstiges ────────────────────────────────────────────────────────

section_title("Topf 2 · Sonstiges (inkl. Termingeschäfte, Dividenden, Zinsen)")

st.markdown(
    '<div class="metric-grid">'
    + metric_card("Dividenden", d['dividends_eur'])
    + metric_card("Zinsen (netto)", d['interest_eur'])
    + (metric_card("Sollzinsen (n. abzf.)", d.get('debit_interest_eur', 0), "info") if abs(d.get('debit_interest_eur', 0)) > 0.01 else '')
    + metric_card("Sonstige Gewinne", d['options_gain_eur'] - adj_cross + (tk_gain_adj.get('Topf2', 0) if tageskurs_aktiv else 0), "gain")
    + metric_card("Sonstige Verluste", d['options_loss_eur'] + (tk_loss_adj.get('Topf2', 0) if tageskurs_aktiv else 0), "loss")
    + metric_card("Saldo Sonstiges", adj_topf_2, "saldo")
    + '</div>',
    unsafe_allow_html=True
)

topf2_cats = d.get('topf2_by_category', {})
if topf2_cats:
    with st.expander("Aufschlüsselung Topf 2"):
        div_eur = d.get('dividends_eur', 0)
        int_eur = d.get('interest_eur', 0)
        cat_table = "| Gattung | Gewinne | Verluste | Netto |\n|---------|--------:|---------:|------:|\n"
        if div_eur >= 0:
            cat_table += f"| Dividenden | {fmt_de(div_eur)} EUR | 0,00 EUR | {fmt_de(div_eur)} EUR |\n"
        else:
            cat_table += f"| Dividenden | 0,00 EUR | {fmt_de(div_eur)} EUR | {fmt_de(div_eur)} EUR |\n"
        if int_eur >= 0:
            cat_table += f"| Zinsen | {fmt_de(int_eur)} EUR | 0,00 EUR | {fmt_de(int_eur)} EUR |\n"
        else:
            cat_table += f"| Zinsen | 0,00 EUR | {fmt_de(int_eur)} EUR | {fmt_de(int_eur)} EUR |\n"
        for cat, vals in sorted(topf2_cats.items()):
            net = vals['gain'] + vals['loss']
            cat_table += f"| {cat} | {fmt_de(vals['gain'])} EUR | {fmt_de(vals['loss'])} EUR | {fmt_de(net)} EUR |\n"
        tk_corr_topf2 = (tk_gain_adj.get('Topf2', 0) + tk_loss_adj.get('Topf2', 0)) if tageskurs_aktiv else 0
        if tageskurs_aktiv and abs(tk_corr_topf2) > 0.01:
            if tk_corr_topf2 >= 0:
                cat_table += f"| Tageskurs-Korrektur | {fmt_de(tk_corr_topf2)} EUR | 0,00 EUR | {fmt_de(tk_corr_topf2)} EUR |\n"
            else:
                cat_table += f"| Tageskurs-Korrektur | 0,00 EUR | {fmt_de(tk_corr_topf2)} EUR | {fmt_de(tk_corr_topf2)} EUR |\n"
        total_gain = sum(v['gain'] for v in topf2_cats.values()) + max(div_eur, 0) + max(int_eur, 0) + (tk_gain_adj.get('Topf2', 0) if tageskurs_aktiv else 0)
        total_loss = sum(v['loss'] for v in topf2_cats.values()) + min(div_eur, 0) + min(int_eur, 0) + (tk_loss_adj.get('Topf2', 0) if tageskurs_aktiv else 0)
        cat_table += f"| **Saldo Topf 2** | **{fmt_de(total_gain)} EUR** | **{fmt_de(total_loss)} EUR** | **{fmt_de(total_gain + total_loss)} EUR** |\n"
        st.markdown(cat_table)

# ── Fremdwährungs-Gewinne/Verluste ──────────────────────────────────────────

fx_results = d.get('fx_results', {})
fx_total_gain = d.get('fx_total_gain', 0)
fx_total_loss = d.get('fx_total_loss', 0)
fx_mtm = d.get('fx_mtm', {})

fx_source = d.get('fx_source', 'none')

if fx_results:
    if fx_source == 'csv':
        section_title("Fremdwährungs-Gewinne/Verluste (IBKR-Bericht)")
    elif fx_source == 'xml':
        section_title("Fremdwährungs-Gewinne/Verluste (XML FxTransactions)")
    else:
        section_title("Fremdwährungs-Gewinne/Verluste (FIFO-Approximation)")

    st.markdown(
        '<div class="metric-grid">'
        + metric_card("FX Gewinne", fx_total_gain, "gain")
        + metric_card("FX Verluste", fx_total_loss, "loss")
        + metric_card("FX Netto", fx_total_gain + fx_total_loss, "saldo")
        + '</div>',
        unsafe_allow_html=True
    )

    with st.expander("Details pro Währung"):
        fx_table = "| Währung | Gewinn | Verlust | Netto | MTM (Vergleich) |\n|---------|--------|---------|-------|----------------|\n"
        for curr, data in sorted(fx_results.items()):
            mtm_val = fx_mtm.get(curr)
            mtm_str = f"{fmt_de(mtm_val)} EUR" if mtm_val is not None else "-"
            fx_table += f"| {curr} | {fmt_de(data['gain'])} | {fmt_de(data['loss'])} | {fmt_de(data['net'])} | {mtm_str} |\n"
        st.markdown(fx_table)
        fx_tgl = d.get('fx_translation', 0)
        if fx_tgl != 0:
            st.markdown(f"**IBKR Referenz (fxTranslationGainLoss):** {fmt_de(fx_tgl)} EUR")

        if fx_source in ('csv', 'xml'):
            st.success("Exakte FX-Werte aus " + ("IBKR Standard-Bericht" if fx_source == 'csv' else "XML FxTransactions") +
                       " (per-Settlement FIFO, alle Währungen).")
        else:
            no_xml_fx = not d.get('xml_has_fx_data', True)
            fx_prior = d.get('fx_has_prior_data', False)
            extra = (" Die Flex Query enthält keine FxTransactions — "
                     "Kursgenauigkeit der Approximation ist eingeschränkt.") if no_xml_fx else ""
            if fx_prior:
                st.warning(f"FIFO-Approximation aus Flex Query (Tagesraten-Substitution).{extra} "
                           "Für exakte Werte: IBKR Standard-Bericht (CSV) oben hochladen.")
            else:
                st.warning(f"**Nur Steuerjahr geladen.** FIFO-Approximation.{extra} "
                           "Für exakte Werte: IBKR Standard-Bericht (CSV) oben hochladen.")
        st.info("**Rechtsgrundlage:** BMF-Schreiben Rn. 131 - verzinsliches Fremdwährungsguthaben, "
                "§20 Abs. 2 S. 1 Nr. 7 EStG (Anlage KAP, Topf 2). FIFO-Methode (§20 Abs. 4 S. 7). "
                "In Topf 2 enthalten.")

# ── Anlage KAP-INV · Investmentfonds (Detail-Anzeige) ─────────────────────────

if has_etf_data and invstg_aktiv:
    etf_gain_raw = kap_inv.get('etf_gain_raw_eur', 0)
    etf_loss_raw = kap_inv.get('etf_loss_raw_eur', 0)
    etf_gain_taxable = kap_inv.get('etf_gain_taxable_eur', 0)
    etf_loss_taxable = kap_inv.get('etf_loss_taxable_eur', 0)
    etf_div_raw = kap_inv.get('etf_dividends_raw_eur', 0)
    etf_div_taxable = kap_inv.get('etf_dividends_taxable_eur', 0)
    etf_wht = kap_inv.get('etf_wht_eur', 0)
    etf_net_taxable = kap_inv.get('etf_net_taxable_eur', 0) + tageskurs_kapinv_corr
    etf_unknown = kap_inv.get('etf_unknown_isins', [])
    etf_stillhalter = kap_inv.get('etf_stillhalter_premium_eur', 0)

if has_etf_data and invstg_aktiv:
    section_title("Anlage KAP-INV · Investmentfonds (InvStG)")

    # Manual classification for unknown ETFs — BEFORE cards so values are correct
    if etf_unknown:
        st.markdown(f"""
<div style="background: rgba(251,191,36,0.08); border: 1px solid rgba(251,191,36,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #fbbf24;">Unbekannte ETFs — manuelle Klassifizierung:</strong> {len(etf_unknown)} ETF(s) nicht in der Klassifizierungstabelle.
    Standardmässig als sonstiger Fonds (0% TFS) behandelt. Falls es sich um Aktienfonds handelt (mind. 51% Aktienquote), bitte unten korrigieren.
</div>
""", unsafe_allow_html=True)
        cls_options = {"Sonstiger Fonds (0% TFS)": 0.0, "Aktienfonds (30% TFS)": 0.30, "Mischfonds (15% TFS)": 0.15}
        for isin in etf_unknown:
            info = etf_by_isin.get(isin, {})
            ticker = info.get('ticker', isin[:12])
            name = info.get('name', '')
            label = f"{ticker} ({isin})" + (f" — {name}" if name else "")
            choice = st.selectbox(label, list(cls_options.keys()), key=f"etf_cls_{isin}")
            new_tfs = cls_options[choice]
            old_tfs = info.get('tfs_rate', 0.0)
            if new_tfs != old_tfs:
                raw_gain = info.get('gain', 0)
                raw_loss = info.get('loss', 0)
                raw_div = info.get('div', 0)
                old_gain_tax = info.get('gain_taxable', raw_gain)
                old_loss_tax = info.get('loss_taxable', raw_loss)
                old_div_tax = info.get('div_taxable', raw_div)
                new_gain_tax = raw_gain * (1 - new_tfs)
                new_loss_tax = raw_loss * (1 - new_tfs)
                new_div_tax = raw_div * (1 - new_tfs)
                etf_gain_taxable += (new_gain_tax - old_gain_tax)
                etf_loss_taxable += (new_loss_tax - old_loss_tax)
                etf_div_taxable += (new_div_tax - old_div_tax)
                etf_net_taxable += (new_gain_tax - old_gain_tax) + (new_loss_tax - old_loss_tax) + (new_div_tax - old_div_tax)
                info['tfs_rate'] = new_tfs
                info['gain_taxable'] = new_gain_tax
                info['loss_taxable'] = new_loss_tax
                info['div_taxable'] = new_div_tax
                cls_map = {0.30: 'aktienfonds', 0.15: 'mischfonds', 0.0: 'sonstiger_fonds'}
                info['classification'] = cls_map.get(new_tfs, 'sonstiger_fonds')

    # Hero card for KAP-INV net taxable
    inv_hero_color = "#4ade80" if etf_net_taxable >= 0 else "#f87171"
    st.markdown(f"""
<div class="hero-card-inv">
    <div class="hero-label">KAP-INV · Steuerpflichtige Erträge (nach Teilfreistellung)</div>
    <div class="hero-value" style="color:{inv_hero_color}">{fmt(etf_net_taxable)}</div>
    <div class="hero-formula">G/V stpfl. ({fmt(etf_gain_taxable + etf_loss_taxable)}) + Div stpfl. ({fmt(etf_div_taxable)}) - QSt ({fmt(etf_wht)})</div>
</div>
""", unsafe_allow_html=True)

    st.markdown(
        '<div class="metric-grid">'
        + metric_card("ETF-Gewinne (roh)", etf_gain_raw, "gain")
        + metric_card("ETF-Verluste (roh)", etf_loss_raw, "loss")
        + metric_card("Teilfreistellung", etf_gain_raw - etf_gain_taxable + etf_loss_raw - etf_loss_taxable, "info")
        + metric_card("ETF-Netto (stpfl.)", etf_gain_taxable + etf_loss_taxable, "saldo")
        + metric_card("ETF-Div. (stpfl.)", etf_div_taxable)
        + metric_card("ETF-Quellensteuer", etf_wht)
        + '</div>',
        unsafe_allow_html=True
    )

    with st.expander("ETF-Details nach ISIN"):
        etf_table = "| Ticker | Typ | TFS | G/V roh | G/V stpfl. | Div roh | Div stpfl. |\n"
        etf_table += "|--------|-----|----:|--------:|----------:|--------:|----------:|\n"
        for isin, info in sorted(etf_by_isin.items(), key=lambda x: x[1].get('ticker', '')):
            cls_short = {'aktienfonds': 'Aktien', 'mischfonds': 'Misch', 'sonstiger_fonds': 'Sonst.'}.get(info.get('classification', ''), '?')
            tfs_pct = f"{info.get('tfs_rate', 0) * 100:.0f}%"
            gv_raw = info.get('gain', 0) + info.get('loss', 0)
            gv_tax = info.get('gain_taxable', 0) + info.get('loss_taxable', 0)
            div_raw = info.get('div', 0)
            div_tax = info.get('div_taxable', 0)
            etf_table += f"| {info.get('ticker', isin)} | {cls_short} | {tfs_pct} | {fmt_de(gv_raw)} | {fmt_de(gv_tax)} | {fmt_de(div_raw)} | {fmt_de(div_tax)} |\n"
        st.markdown(etf_table)


# ── Anlage SO · Private Veräußerungsgeschäfte (§23 EStG) ──────────────────────

anlage_so = d.get('anlage_so', {})
so_by_isin = anlage_so.get('by_isin', {})
has_so_data = bool(so_by_isin)

if has_so_data:
    section_title("Anlage SO · Private Veräußerungsgeschäfte (§23 EStG)")

    so_taxable = anlage_so.get('taxable_gain', 0) + anlage_so.get('taxable_loss', 0)
    so_free = anlage_so.get('tax_free_gain', 0) + anlage_so.get('tax_free_loss', 0)
    so_total = anlage_so.get('total_gain', 0) + anlage_so.get('total_loss', 0)

    so_unknown = abs(anlage_so.get('unknown_gain', 0)) + abs(anlage_so.get('unknown_loss', 0))
    so_has_unknown = so_unknown > 0.01

    history_hint = ""
    if so_has_unknown:
        history_hint = (
            '<br><br><strong style="color: #f87171;">Haltedauer nicht ermittelbar:</strong> '
            'Ohne Vorjahres-XMLs oder CLOSED_LOT-Daten kann die Haltedauer nicht bestimmt werden. '
            f'Betroffene Positionen ({fmt(so_unknown)}) werden konservativ als <strong>steuerpflichtig</strong> behandelt. '
            'Laden Sie die XML des Kaufjahres als History-Datei hoch, um die Spekulationsfrist korrekt zu berechnen.'
        )

    st.markdown(f"""
<div style="background: rgba(251,191,36,0.08); border: 1px solid rgba(251,191,36,0.25); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.8rem; color: #94a3b8;">
    <strong style="color: #fbbf24;">Physische Gold-ETCs mit Lieferanspruch</strong> werden nach
    <strong>§23 Abs. 1 S. 1 Nr. 2 EStG</strong> als private Veräußerungsgeschäfte behandelt
    (bestätigt durch <strong>BFH VIII R 4/15</strong> für Xetra-Gold).
    Die Spekulationsfrist beträgt <strong>1 Jahr</strong> (§23 Abs. 1 S. 1 Nr. 2 S. 1 EStG) -
    Veräußerungsgewinne nach Ablauf dieser Frist sind <strong style="color: #4ade80;">steuerfrei</strong>.
    Innerhalb der Frist sind sie auf <strong>Anlage SO</strong> zu erklären (nicht auf Anlage KAP).{history_hint}
</div>
""", unsafe_allow_html=True)

    so_hero_color = "#fbbf24" if so_taxable >= 0 else "#f87171"
    st.markdown(f"""
<div class="hero-card-so">
    <div class="hero-label">Anlage SO · Steuerpflichtiger Gewinn/Verlust (Haltedauer ≤ 1 Jahr)</div>
    <div class="hero-value" style="color:{so_hero_color}">{fmt(so_taxable)}</div>
    <div class="hero-formula">Gesamt {fmt(so_total)} − Steuerfrei {fmt(so_free)} (Haltedauer > 1 Jahr)</div>
</div>
""", unsafe_allow_html=True)

    st.markdown(
        '<div class="metric-grid">'
        + metric_card("Gesamt G/V", so_total, "saldo")
        + metric_card("Steuerfrei (> 1J)", so_free, "gain")
        + metric_card("Steuerpflichtig (≤ 1J)", so_taxable, "loss" if so_taxable < 0 else "info")
        + '</div>',
        unsafe_allow_html=True
    )

    with st.expander("Gold-ETC Details nach ISIN"):
        so_table = "| Ticker | Gesamt | Steuerfrei (> 1J) | Steuerpflichtig (≤ 1J) |\n"
        so_table += "|--------|-------:|------------------:|-----------------------:|\n"
        for isin, info in sorted(so_by_isin.items(), key=lambda x: abs(x[1]['total']), reverse=True):
            so_table += f"| {info.get('ticker', isin)} | {fmt_de(info['total'])} | {fmt_de(info.get('tax_free', 0))} | {fmt_de(info.get('taxable', 0))} |\n"
        st.markdown(so_table)

        so_details = anlage_so.get('details', [])
        if so_details:
            st.markdown("**Lot-Details (FIFO-Zuordnung):**")
            lot_table = "| Ticker | Kauf | Verkauf | Stk. | G/V (EUR) | Status |\n"
            lot_table += "|--------|------|---------|-----:|----------:|--------|\n"
            for lot in so_details:
                status = "steuerfrei" if lot.get('is_tax_free') else "steuerpflichtig"
                qty = abs(lot.get('quantity', 0))
                lot_table += f"| {lot['ticker']} | {lot.get('open_date', '?')} | {lot.get('close_date', '?')} | {qty:.0f} | {fmt_de(lot['pnl_eur'])} | {status} |\n"
            st.markdown(lot_table)


# ── Einzelnachweise · Trade-Level-Reporting ──────────────────────────────────
trade_details = list(d.get('trade_details', []))  # copy to avoid mutating report_data

# Inject per-lot FX corrections when Tageskurs is active
if trade_details and tageskurs_aktiv:
    for lot in d.get('fx_correction_details', []):
        if abs(lot.get('delta_eur', 0)) < 0.005:
            continue
        underlying = (lot.get('underlyingSymbol', '') or lot.get('symbol', '') or '').split()[0]
        open_dt = (lot.get('openDateTime', '') or '')[:10]
        close_dt = lot.get('reportDate', '')
        trade_details.append({
            'dateTime': close_dt, 'reportDate': close_dt,
            'symbol': lot.get('symbol', ''),
            'description': f'Tageskurs-Korrektur (Kauf {open_dt}, Kurs {lot["fx_open"]:.5f} → {lot["fx_close"]:.5f})',
            'isin': lot.get('isin', ''),
            'assetCategory': lot.get('assetCategory', ''),
            'subCategory': lot.get('subCategory', ''),
            'buySell': '', 'openClose': '',
            'quantity': lot.get('quantity', ''),
            'transactionType': 'FX-Korrektur', 'currency': lot.get('currency', ''),
            'tradePrice': 0, 'cost': lot.get('cost', 0), 'proceeds': 0,
            'fifoPnlRealized': 0, 'fxRateToBase': 0,
            'pnl_eur': lot['delta_eur'],
            'topf': lot.get('topf', 'Topf2'),
            'strike': '', 'expiry': '', 'putCall': '', 'multiplier': '',
            'underlyingSymbol': underlying,
            'source': 'tageskurs_korrektur',
        })
    trade_details.sort(key=lambda r: r.get('dateTime', '') or r.get('reportDate', '') or 'zzzz')

if trade_details:
    section_title('Einzelnachweise - Trade-Details (Excel)')
    st.markdown('<div style="margin-top:-18px;margin-bottom:12px"><span style="background:linear-gradient(135deg,#f59e0b,#ef4444);color:#fff;padding:5px 16px;border-radius:6px;font-size:0.85em;font-weight:700;letter-spacing:1px;box-shadow:0 2px 8px rgba(239,68,68,0.3)">NEUE FUNKTION &middot; BETA</span></div>', unsafe_allow_html=True)

    from collections import defaultdict

    topf_readable = {
        'Topf1': 'Topf 1 - Aktien (§20 Abs. 2 Nr. 1 EStG)',
        'Topf2': 'Topf 2 - Sonstiges (Termingeschäfte, Stillhalter, FX)',
        'KAP-INV': 'Anlage KAP-INV (InvStG)',
        'Anlage SO': 'Anlage SO (§23 EStG)',
    }
    cat_labels = {
        'STK': 'Aktie', 'OPT': 'Option', 'FUT': 'Future',
        'FOP': 'Futures-Option', 'FSFOP': 'Flex-Option',
        'BILL': 'T-Bill', 'BOND': 'Anleihe',
    }

    trades_by_topf = defaultdict(list)
    for row in trade_details:
        trades_by_topf[row.get('topf', 'Topf2')].append(row)

    # Kurzübersicht
    n_trades = sum(1 for r in trade_details if r.get('source') == 'trades')
    n_korr = len(trade_details) - n_trades
    n_underlyings = len(set(
        (r.get('underlyingSymbol', '') or r.get('symbol', '') or '?').split()[0]
        for r in trade_details if r.get('source') == 'trades'
    ))
    header = f"**{n_trades} Trades, {n_underlyings} Wertpapiere"
    if n_korr > 0:
        header += f" (+ {n_korr} Korrekturen/Zuflüsse)"
    header += "**"
    summary_lines = [header]
    for topf_key in ['Topf1', 'Topf2', 'KAP-INV', 'Anlage SO']:
        rows = trades_by_topf.get(topf_key, [])
        if rows:
            s = sum(r.get('pnl_eur', 0) for r in rows)
            label = topf_readable.get(topf_key, topf_key).split(' - ')[0]
            summary_lines.append(f"{label}: {fmt_de(s)} EUR")
    st.markdown(" | ".join(summary_lines))
    st.caption("Enthält Trades, Optionsverläufe, Stillhalter-Korrekturen und Zuflüsse. Dividenden, Zinsen und Quellensteuer sind nicht enthalten (siehe IBKR-Kontoauszug).")

    def _format_instrument(row):
        sym = row.get('symbol', '') or ''
        desc = row.get('description', '') or ''
        pc = row.get('putCall', '') or ''
        strike = row.get('strike', '') or ''
        expiry = row.get('expiry', '') or ''
        if pc and strike and expiry:
            pc_label = 'Call' if pc == 'C' else 'Put'
            if len(expiry) == 8:
                exp_fmt = f"{expiry[:4]}-{expiry[4:6]}-{expiry[6:]}"
            else:
                exp_fmt = expiry[:10]
            return f"{sym} ({pc_label} {strike} exp. {exp_fmt})"
        if desc and sym:
            return f"{sym} ({desc})"
        return sym or desc or ''

    def _get_group_key(row):
        us = (row.get('underlyingSymbol', '') or '').strip()
        if us:
            return us.split()[0]
        sym = (row.get('symbol', '') or '').strip()
        return sym.split()[0] if sym else '?'

    def _build_excel(trade_details, trades_by_topf):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = f"Trade-Details {steuerjahr}"

        # Styles
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="1e3a5f")
        grp_font = Font(bold=True, size=10)
        grp_fill = PatternFill("solid", fgColor="d6e4f0")
        gain_font = Font(color="006100", size=9)
        gain_fill = PatternFill("solid", fgColor="e2efda")
        loss_font = Font(color="9c0006", size=9)
        loss_fill = PatternFill("solid", fgColor="fce4ec")
        korr_font = Font(italic=True, size=9)
        korr_fill = PatternFill("solid", fgColor="fff9c4")
        sub_font = Font(bold=True, size=9)
        sub_fill = PatternFill("solid", fgColor="f2f2f2")
        total_font = Font(bold=True, size=10, color="FFFFFF")
        total_fill = PatternFill("solid", fgColor="4a4a4a")
        normal_font = Font(size=9)
        thin_border = Border(bottom=Side(style='thin', color='cccccc'))
        num_fmt_eur = '#.##0,00'
        num_fmt_4d = '#.##0,0000'

        cols = ['Datum', 'Handelsdatum', 'Wertpapier', 'ISIN', 'Kategorie',
                'K/V', 'Stk.', 'Kurs', 'Kostenbasis', 'Erlöse',
                'G/V (Orig.)', 'Kommission', 'Währung', 'Wechselkurs', 'G/V (EUR)', 'Anmerkung']
        col_widths = [12, 12, 42, 15, 10, 6, 8, 11, 13, 13, 13, 11, 6, 11, 14, 40]
        eur_col = 15  # 1-based: column O = G/V (EUR)

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        row_num = 1
        grand_total = 0.0

        for topf_key in ['Topf1', 'Topf2', 'KAP-INV', 'Anlage SO']:
            topf_rows = trades_by_topf.get(topf_key, [])
            if not topf_rows:
                continue

            # Topf Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(cols))
            cell = ws.cell(row=row_num, column=1, value=topf_readable.get(topf_key, topf_key))
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='left')
            row_num += 1

            # Column headers
            for ci, col_name in enumerate(cols, 1):
                cell = ws.cell(row=row_num, column=ci, value=col_name)
                cell.font = Font(bold=True, size=9)
                cell.fill = PatternFill("solid", fgColor="e8e8e8")
                cell.border = thin_border
            row_num += 1

            # Group by underlying
            groups = defaultdict(list)
            for r in topf_rows:
                groups[_get_group_key(r)].append(r)

            topf_total = 0.0
            for grp_key in sorted(groups.keys()):
                grp_rows = groups[grp_key]
                grp_rows.sort(key=lambda r: r.get('dateTime', '') or r.get('reportDate', '') or '')

                # Find description for group header
                grp_desc = ''
                grp_isin = ''
                for r in grp_rows:
                    if r.get('description') and r.get('source') != 'stillhalter_korrektur':
                        grp_desc = r['description']
                    if r.get('isin'):
                        grp_isin = r['isin']
                    if grp_desc and grp_isin:
                        break
                grp_label = f"{grp_key}"
                if grp_desc:
                    grp_label += f" - {grp_desc}"
                if grp_isin:
                    grp_label += f" ({grp_isin})"

                # Group header row
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(cols))
                cell = ws.cell(row=row_num, column=1, value=grp_label)
                cell.font = grp_font
                cell.fill = grp_fill
                row_num += 1

                grp_has_stillhalter = any(
                    r.get('source') in ('stillhalter_korrektur', 'cross_year_put_korrektur')
                    for r in grp_rows
                )

                grp_total = 0.0
                for r in grp_rows:
                    source = r.get('source', '')
                    pnl_eur = r.get('pnl_eur', 0)
                    pnl_orig = r.get('fifoPnlRealized', 0)
                    fx = r.get('fxRateToBase', 0)
                    cost = r.get('cost', 0)
                    proceeds = r.get('proceeds', 0)
                    price = r.get('tradePrice', 0)
                    mult = r.get('multiplier', '')

                    anmerkung = ''
                    if source == 'pnl_summary':
                        anmerkung = 'Aus IBKR PnL-Summary'
                    elif source == 'stillhalter_korrektur':
                        anmerkung = r.get('description', 'Korrektur')
                    elif source == 'zufluss':
                        anmerkung = r.get('description', 'Zufluss §11 EStG')
                    elif source == 'zufluss_korrektur':
                        anmerkung = r.get('description', 'Vorjahres-Korrektur')
                    elif source == 'tageskurs_korrektur':
                        anmerkung = r.get('description', 'Tageskurs §20 Abs. 4')
                    elif source == 'cross_year_put_korrektur':
                        anmerkung = r.get('description', 'Cross-Year Put-Korrektur')
                    elif source == 'trades' and grp_has_stillhalter and r.get('assetCategory') == 'STK':
                        anmerkung = 'Kostenbasis enthält Stillhalterprämie (s. Korrekturzeile)'

                    # Readable open/close label
                    bs = r.get('buySell', '')
                    oc = r.get('openClose', '')
                    if bs == 'SELL' and oc == 'O':
                        bs_label = 'STO'
                    elif bs == 'BUY' and oc == 'C':
                        bs_label = 'BTC'
                    elif bs == 'BUY' and oc == 'O':
                        bs_label = 'BTO'
                    elif bs == 'SELL' and oc == 'C':
                        bs_label = 'STC'
                    else:
                        bs_label = bs

                    commission = r.get('ibCommission', 0) or 0

                    values = [
                        (r.get('reportDate', '') or '')[:10],
                        (r.get('dateTime', '') or '')[:10],
                        _format_instrument(r),
                        r.get('isin', ''),
                        cat_labels.get(r.get('assetCategory', ''), r.get('assetCategory', '')),
                        bs_label,
                        r.get('quantity', ''),
                        price if price else None,
                        cost if cost else None,
                        proceeds if proceeds else None,
                        pnl_orig if pnl_orig else None,
                        commission if commission else None,
                        r.get('currency', ''),
                        fx if fx else None,
                        pnl_eur,
                        anmerkung,
                    ]

                    for ci, val in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=ci, value=val)
                        cell.font = normal_font
                        # Number formats
                        if ci in (8, 9, 10, 11, 12, 15) and isinstance(val, (int, float)):
                            cell.number_format = num_fmt_eur
                        elif ci == 14 and isinstance(val, (int, float)):
                            cell.number_format = num_fmt_4d

                    # Row coloring
                    if source in ('stillhalter_korrektur', 'zufluss', 'zufluss_korrektur', 'tageskurs_korrektur', 'cross_year_put_korrektur'):
                        for ci in range(1, len(cols) + 1):
                            ws.cell(row=row_num, column=ci).fill = korr_fill
                            ws.cell(row=row_num, column=ci).font = korr_font
                    elif pnl_eur > 0.005:
                        for ci in range(1, len(cols) + 1):
                            ws.cell(row=row_num, column=ci).fill = gain_fill
                            ws.cell(row=row_num, column=ci).font = gain_font
                    elif pnl_eur < -0.005:
                        for ci in range(1, len(cols) + 1):
                            ws.cell(row=row_num, column=ci).fill = loss_fill
                            ws.cell(row=row_num, column=ci).font = loss_font

                    grp_total += pnl_eur
                    row_num += 1

                # Group subtotal
                ws.cell(row=row_num, column=1, value=f"Zwischensumme {grp_key}")
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=row_num, column=ci).font = sub_font
                    ws.cell(row=row_num, column=ci).fill = sub_fill
                    ws.cell(row=row_num, column=ci).border = thin_border
                cell = ws.cell(row=row_num, column=eur_col, value=grp_total)
                cell.number_format = num_fmt_eur
                cell.font = sub_font
                cell.fill = sub_fill
                topf_total += grp_total
                row_num += 1

            # Topf total
            topf_label = topf_readable.get(topf_key, topf_key).split(' - ')[0]
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=eur_col - 1)
            cell = ws.cell(row=row_num, column=1, value=f"SUMME {topf_label}")
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal='right')
            for ci in range(1, len(cols) + 1):
                ws.cell(row=row_num, column=ci).fill = total_fill
            cell = ws.cell(row=row_num, column=eur_col, value=topf_total)
            cell.number_format = num_fmt_eur
            cell.font = total_font
            cell.fill = total_fill
            grand_total += topf_total
            row_num += 2  # blank row between topf sections

        # Freeze header: first 2 rows are topf header + column headers,
        # but since they repeat, freeze after row 1
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    xlsx_data = _build_excel(trade_details, trades_by_topf)

    st.download_button(
        label=f"Trade-Details als Excel herunterladen ({len(trade_details)} Positionen)",
        data=xlsx_data,
        file_name=f"trade_details_{steuerjahr}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


# ── Plausibilitätscheck (wenn CSV hochgeladen) ─────────────────────────────
csv_cats = d.get('csv_category_totals', {})
if csv_cats:
    section_title("Plausibilitätscheck (IBKR-Bericht vs. Berechnung)")
    # Map IBKR categories to our Topf structure
    # Add back stillhalter premium AND cross-year put correction for IBKR comparison
    # (IBKR doesn't separate these — we do, so we reverse for the plausibility check)
    cross_put = d['audit'].get('cross_year_put_total', 0)
    no_invstg_gain = d['audit'].get('no_invstg_gain', 0)
    no_invstg_loss = d['audit'].get('no_invstg_loss', 0)
    # Add back premiums that WERE subtracted from stocks_gain, but NOT put premiums
    # where the stock wasn't sold (those were never in stocks_gain to begin with)
    stillhalter_addback = d['audit'].get('stillhalter_premium_eur', 0) - d['audit'].get('put_nosell_premium_eur', 0)
    our_stk_gain = d['stocks_gain_eur'] + stillhalter_addback + cross_put + no_invstg_gain
    our_stk_loss = d['stocks_loss_eur'] + no_invstg_loss

    # If InvStG active: add ETF values back for IBKR comparison (IBKR counts ETFs as Aktien)
    # Note: stillhalter_premium_eur already includes the ETF portion — don't add etf_stillhalter again
    kap_inv_data = d.get('kap_inv', {})
    if has_etf_data and invstg_aktiv:
        our_stk_gain += kap_inv_data.get('etf_gain_raw_eur', 0)
        our_stk_loss += kap_inv_data.get('etf_loss_raw_eur', 0)

    # Add back Anlage SO Gold-ETC values for IBKR comparison (IBKR counts them as STK)
    if has_so_data:
        our_stk_gain += anlage_so.get('total_gain', 0)
        our_stk_loss += anlage_so.get('total_loss', 0)

    ibkr_topf2_cats = ["Aktien- und Indexoptionen", "Futures", "Optionen auf Futures (Future-Style)",
                        "Optionen auf Futures", "Anleihen", "Treasury Bills"]
    # Reverse Zufluss adjustments for IBKR comparison (IBKR doesn't know about Zufluss)
    zufluss_adj = d['audit'].get('zufluss_premium_eur', 0) - d['audit'].get('prior_zufluss_correction_eur', 0)
    our_topf2_gain = d['options_gain_eur'] - d['audit'].get('stillhalter_premium_eur', 0) - d.get('fx_total_gain', 0) - no_invstg_gain - zufluss_adj
    our_topf2_loss = d['options_loss_eur'] - d.get('fx_total_loss', 0) - no_invstg_loss
    ibkr_topf2_gain = sum(csv_cats.get(c, {}).get('gain', 0) for c in ibkr_topf2_cats)
    ibkr_topf2_loss = sum(csv_cats.get(c, {}).get('loss', 0) for c in ibkr_topf2_cats)

    ibkr_stk = csv_cats.get('Aktien', {})
    ibkr_fx = csv_cats.get('Devisen', {})

    # Dividenden/Quellensteuer: ETF-Werte für IBKR-Vergleich zurückaddieren
    our_div = d['dividends_eur']
    our_wht = d['withholding_tax_eur']
    if has_etf_data and invstg_aktiv:
        our_div += kap_inv_data.get('etf_dividends_raw_eur', 0)
        our_wht += kap_inv_data.get('etf_wht_eur', 0)

    rows = [
        ("Aktien (Topf 1) Netto", ibkr_stk.get('net', 0), our_stk_gain + our_stk_loss),
        ("Sonstiges (Topf 2) Netto", ibkr_topf2_gain + ibkr_topf2_loss, our_topf2_gain + our_topf2_loss),
        ("FX (Devisen) Netto", ibkr_fx.get('net', 0), fx_total_gain + fx_total_loss),
    ]

    # Dividenden, Zinsen, Quellensteuer aus CSV
    csv_income = d.get('csv_income_totals', {})
    if 'dividends_eur' in csv_income:
        rows.append(("Dividenden", csv_income['dividends_eur'], our_div))
    if 'interest_eur' in csv_income:
        # IBKR-Bericht enthält Sollzinsen (DINT) — für Vergleich zurückaddieren
        our_interest_for_comparison = d['interest_eur'] + d.get('debit_interest_eur', 0)
        rows.append(("Zinsen", csv_income['interest_eur'], our_interest_for_comparison))
    if 'withholding_tax_eur' in csv_income:
        rows.append(("Quellensteuer", abs(csv_income['withholding_tax_eur']), our_wht))

    check_table = "| Kategorie | IBKR-Bericht | Unsere Berechnung | Differenz |\n|-----------|-------------|-------------------|----------|\n"
    all_match = True
    zinsen_fx_diff = False
    for label, ibkr_val, our_val in rows:
        diff = our_val - ibkr_val
        match = abs(diff) < 1.0
        if not match:
            if label == "Zinsen":
                zinsen_fx_diff = True
            else:
                all_match = False
        icon = "" if match else (" **(FX)**" if label == "Zinsen" else " **(!)**")
        check_table += f"| {label} | {fmt_de(ibkr_val)} | {fmt_de(our_val)} | {fmt_de(diff)}{icon} |\n"
    st.markdown(check_table)
    if all_match and not zinsen_fx_diff:
        st.success("Alle Kategorien stimmen mit dem IBKR-Bericht überein.")
    elif all_match and zinsen_fx_diff:
        st.success("Alle Kategorien stimmen überein. Zinsen-Differenz ist eine bekannte FX-Konvertierungsdifferenz "
                   "(IBKR konvertiert Fremdwährungs-Anleiheposten im CSV-Bericht mit anderen Kursen als in der XML-BaseCurrency-Ansicht).")
    else:
        st.info("Kleine Abweichungen sind normal (FX-Rundung, Steuerkorrekturen aus Vorjahren).")
    if has_etf_data and invstg_aktiv:
        st.caption("InvStG aktiv: ETF-Werte wurden für diesen Vergleich zurückaddiert, da der IBKR-Bericht keine InvStG-Trennung kennt.")
    if has_so_data:
        st.caption("Anlage SO aktiv: Gold-ETC-Werte wurden für diesen Vergleich zurückaddiert, da IBKR sie als Aktien zählt.")
    if tageskurs_aktiv or zufluss_adj != 0:
        notes = []
        if tageskurs_aktiv:
            corr_sign = "+" if fx_corr_total >= 0 else ""
            notes.append(f"Tageskurs-Korrektur ({corr_sign}{fmt_de(fx_corr_total)} EUR)")
        if zufluss_adj != 0:
            notes.append(f"Stillhalter-Zufluss ({'+' if zufluss_adj >= 0 else ''}{fmt_de(zufluss_adj)} EUR)")
        excluded = " und ".join(notes)
        st.caption(
            f"Der Plausibilitätscheck vergleicht unsere Berechnung 1:1 gegen IBKR's eigene Summen. "
            f"Steuerliche Korrekturen, die über IBKR's Zahlen hinausgehen, werden dabei herausgerechnet: "
            f"{excluded}. So lässt sich prüfen, ob die Basisdaten korrekt verarbeitet wurden, "
            f"bevor die steuerlichen Anpassungen darauf aufsetzen."
        )

# ── Anlage KAP Zeilen ────────────────────────────────────────────────────────

section_title(f"Anlage KAP {steuerjahr} · Eintragungen")

kap_rows_html = (
    kap_row("Z. 19", "Ausländische Kapitalerträge (Netto)", adj_zeile_19, highlight=True)
    + kap_row("Z. 20", "Davon: Aktiengewinne", zeile_20)
    + kap_row("Z. 22", "Verluste ohne Aktien", zeile_22, force_positive=True)
    + kap_row("Z. 23", "Aktienverluste", zeile_23, force_positive=True)
    + kap_row("Z. 41", "Anrechenbare Quellensteuer", quellensteuer)
)

if has_etf_data and invstg_aktiv:
    kap_rows_html += '<div class="section-title" style="margin-top:1.5rem;">Anlage KAP-INV</div>'
    kap_rows_html += kap_row("KAP-INV", "Erträge nach Teilfreistellung", etf_net_taxable, highlight=True)
    kap_rows_html += kap_row("KAP-INV", "Anrechenbare Quellensteuer (ETF)", etf_wht)

if has_so_data:
    so_taxable_for_row = anlage_so.get('taxable_gain', 0) + anlage_so.get('taxable_loss', 0)
    so_free_for_row = anlage_so.get('tax_free_gain', 0) + anlage_so.get('tax_free_loss', 0)
    kap_rows_html += '<div class="section-title" style="margin-top:1.5rem;">Anlage SO (§23 EStG)</div>'
    kap_rows_html += kap_row("SO", "Steuerpflichtiger Gewinn/Verlust (≤ 1 Jahr)", so_taxable_for_row, highlight=True)
    if abs(so_free_for_row) > 0.01:
        kap_rows_html += kap_row("SO", "Steuerfrei (> 1 Jahr Haltedauer)", so_free_for_row)

st.markdown(kap_rows_html, unsafe_allow_html=True)

# ── Multi-Account Breakdown ─────────────────────────────────────────────────

if n_accounts > 1:
    with st.expander(f"Aufschlüsselung nach Konten ({n_accounts} Konten)"):
        acct_table = "| Konto | Topf 1 (Aktien) | Topf 2 (Sonstiges) | Z. 19 (Netto) | Z. 41 (QSt) |\n"
        acct_table += "|-------|----------------:|-------------------:|--------------:|------------:|\n"
        for idx, (name, rep) in enumerate(zip(account_names, reports)):
            t1 = rep.get('topf_1_aktien_netto', 0)
            t2 = rep.get('topf_2_sonstiges_netto', 0)
            z19 = rep.get('zeile_19_netto_eur', t1 + t2)
            z41 = rep.get('withholding_tax_eur', 0)
            label = f"Konto {idx+1} ({name})"
            acct_table += f"| {label} | {fmt_de(t1)} | {fmt_de(t2)} | {fmt_de(z19)} | {fmt_de(z41)} |\n"
        acct_table += f"| **Gesamt** | **{fmt_de(topf_1)}** | **{fmt_de(adj_topf_2)}** | **{fmt_de(adj_zeile_19)}** | **{fmt_de(quellensteuer)}** |\n"
        st.markdown(acct_table)
        st.info("Jedes Konto wurde vollständig separat berechnet (eigene Trades, Dividenden, FX-Berechnung, "
                "Stillhalter-Erkennung). Die Einzelergebnisse wurden anschließend addiert.")

# ── Zuflussprinzip Details ────────────────────────────────────────────────────

if zuflussprinzip_aktiv and cross_year_details:
    section_title("Zuflussprinzip · Vorjahres-Prämien (BMF Rn. 25, 33)")
    st.markdown("""
<div style="background: rgba(168,85,247,0.06); border: 1px solid rgba(168,85,247,0.2); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.82rem; color: #94a3b8;">
    Die folgenden Stillhalterprämien wurden <strong>aus dem aktuellen Steuerjahr herausgerechnet</strong>,
    da der Zufluss (= Verkauf der Option) in einem Vorjahr stattfand.
    Diese Beträge gehören in die <strong>Steuererklärung des jeweiligen Vorjahres</strong>.
</div>
""", unsafe_allow_html=True)

    detail_table = "| Symbol | Strike | Verkauf (Zufluss) | Assignment | Prämie (EUR) |\n"
    detail_table += "|--------|--------|-------------------|------------|-------------:|\n"
    for det in cross_year_details:
        detail_table += (f"| {det['symbol']} | {det['strike']} | "
                        f"{det['orig_sell_date']} | "
                        f"{det['assignment_date']} | "
                        f"{fmt_de(det['premium_eur'])} |\n")
    st.markdown(detail_table)

    st.markdown("**Zusammenfassung nach Zuflussjahr:**")
    year_table = "| Steuerjahr | Prämien-Summe (EUR) | Hinweis |\n"
    year_table += "|:----------:|--------------------:|--------|\n"
    for year in sorted(cross_year_by_year.keys()):
        year_table += f"| {year} | {fmt_de(cross_year_by_year[year])} | In Steuererklärung {year} eintragen |\n"
    st.markdown(year_table)

    st.info(f"**Gesamtbetrag Vorjahres-Prämien:** {fmt_de(cross_year_premium)} EUR — "
            f"um diesen Betrag wurde Zeile 19 im aktuellen Jahr reduziert.")

# ── Stillhalter-Zufluss Details (offene Positionen + Vorjahres-Korrekturen) ──

if zufluss_details or prior_zufluss_details:
    section_title("Stillhalter-Zufluss · Offene Positionen & Korrekturen")
    if zufluss_details:
        st.markdown(f"""
<div style="background: rgba(34,197,94,0.06); border: 1px solid rgba(34,197,94,0.2); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.82rem; color: #94a3b8;">
    <strong style="color: #22c55e;">Stillhalter-Zufluss (§11 EStG):</strong> {len(zufluss_details)} Short-Option(en) im Steuerjahr verkauft, deren Position am Jahresende noch offen ist.
    Die Prämien ({fmt_de(zufluss_premium)} EUR) sind als Zufluss im Steuerjahr steuerpflichtig und wurden zu Topf 2 addiert.
</div>
""", unsafe_allow_html=True)
        zt = "| Symbol | Verkaufsdatum | Stk. | Prämie (EUR) |\n"
        zt += "|--------|--------------|-----:|-------------:|\n"
        for det in zufluss_details:
            zt += f"| {det['symbol']} | {det['sell_date'][:10]} | {det['quantity']} | {fmt_de(det['premium_eur'])} |\n"
        st.markdown(zt)

    if prior_zufluss_details:
        st.markdown(f"""
<div style="background: rgba(168,85,247,0.06); border: 1px solid rgba(168,85,247,0.2); border-radius: 10px; padding: 0.75rem 1rem; margin-bottom: 1rem; font-size: 0.82rem; color: #94a3b8;">
    <strong style="color: #a855f7;">Vorjahres-Korrektur:</strong> {len(prior_zufluss_details)} Position(en) wurden in einem Vorjahr als Stillhalter eröffnet und im Steuerjahr glattgestellt.
    Die Prämien ({fmt_de(prior_zufluss_correction)} EUR) waren bereits im Verkaufsjahr steuerpflichtig und wurden vom aktuellen PnL abgezogen.
</div>
""", unsafe_allow_html=True)
        pt = "| Symbol | Verkaufsjahr | Stk. | Korrektur (EUR) |\n"
        pt += "|--------|:-----------:|-----:|----------------:|\n"
        for det in prior_zufluss_details:
            pt += f"| {det['symbol']} | {det['sell_year']} | {det['quantity']} | -{fmt_de(det['premium_eur'])} |\n"
        st.markdown(pt)

# ── Steuerliche Regeln & Berechnungsmethodik ─────────────────────────────────

section_title("Steuerliche Regeln & Berechnungsmethodik")

sh_count = audit.get('stillhalter_count', 0)
sh_eur = audit.get('stillhalter_premium_eur', 0)
base_curr = d.get('base_currency', 'USD')

with st.expander("Regeln anzeigen  - So kommen die Ergebnisse zustande"):
    st.markdown(f"""
### Zwei-Töpfe-Struktur (§20 Abs. 6 EStG)

Das deutsche Steuerrecht unterscheidet zwei getrennte Verrechnungstöpfe:

| Topf | Inhalt | Verlustverrechnung |
|------|--------|-------------------|
| **Topf 1  - Aktien** | Gewinne und Verluste aus Aktienveräußerungen (STK) | Aktienverluste dürfen **nur** mit Aktiengewinnen verrechnet werden. Überschüsse werden als Verlustvortrag ins nächste Jahr übertragen. |
| **Topf 2  - Sonstiges** | Optionen, Futures, Anleihen, T-Bills, Dividenden, Zinsen, Stillhalterprämien | Alle Verluste frei mit allen Gewinnen in Topf 2 verrechenbar. |

Das Finanzamt wendet die Verlustverrechnungsbeschränkung anhand der Zeilen 20 und 23 an  - der Steuerpflichtige meldet die Bruttowerte.

---

### Instrument-Klassifikation

| IBKR-Kategorie | Steuerliche Einordnung | Topf |
|----------------|----------------------|------|
| **STK** (Aktien) | Aktienveräußerung (§20 Abs. 2 Nr. 1) | Topf 1 |
| **OPT** (Optionen) | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **FUT** (Futures) | Termingeschäft  - Festgeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **FOP** (Futures-Optionen) | Termingeschäft (§20 Abs. 2 Nr. 3) | Topf 2 |
| **BILL** (T-Bills) | Kapitalforderung (§20 Abs. 2 Nr. 7) | Topf 2 |
| **BOND** (Anleihen) | Kapitalforderung (§20 Abs. 2 Nr. 7) | Topf 2 |
| **DIV/PIL** (Dividenden) | Laufende Erträge (§20 Abs. 1 Nr. 1) | Topf 2 |
| **INTR/CINT** (Zinsen) | Zinserträge (§20 Abs. 1 Nr. 7) | Topf 2 |
| **CASH/FOREX** (Fremdwährung) | Verzinsl. Fremdwährungsguthaben (§20 Abs. 2 Nr. 7, BMF Rn. 131) | Topf 2 |

---

### Fremdwährungs-Gewinne/Verluste (BMF Rn. 131)

Beim Halten von Fremdwährungsguthaben (z.B. USD) auf einem verzinslichen Konto (IBKR zahlt Zinsen) entstehen bei Kursänderungen steuerlich relevante Gewinne oder Verluste:

- **Anschaffung** = jeder Zufluss von Fremdwährung (Kauf, Dividende, Verkaufserlös)
- **Veräußerung** = jeder Abfluss (Rücktausch, Aktienkauf, Gebühren)
- **FIFO-Methode**: die zuerst erworbenen Beträge werden zuerst veräußert
- **Rechtsgrundlage**: §20 Abs. 2 S. 1 Nr. 7 EStG, Anlage KAP, Topf 2

**Hinweis:** Ohne Vorjahres-XMLs wird der Jahresanfangsbestand zum 01.01.-Kurs als Anschaffung angesetzt (Vereinfachung). Für exakte Berechnung können Vorjahres-XMLs hochgeladen werden. IBKR liefert keine FIFO-Daten für Währungsgewinne, diese werden hier eigenständig berechnet.

---

### Stillhalterprämien bei Assignments (BMF Rn. 25–35)

Wenn Sie eine Option verkaufen (Stillhalter) und diese ausgeübt wird (Assignment), muss die Prämie steuerlich korrekt zugeordnet werden:

- **Prämie** = laufende Einnahmen nach §20 Abs. 1 Nr. 11 → gehört in **Topf 2**
- **Aktientransaktion** = Veräußerung (Call, Rn. 26) bzw. Anschaffung (Put, Rn. 33) nach §20 Abs. 2 → gehört in **Topf 1**

Bei **beiden** Assignment-Typen gilt laut BMF: „Die vereinnahmte Optionsprämie wird bei der Ermittlung des Veräußerungsgewinns **nicht berücksichtigt**." IBKR bündelt die Prämie jedoch im Aktien-Trade (Call: im Verkaufserlös, Put: in den reduzierten Anschaffungskosten). Dieses Tool erkennt Assignments automatisch und trennt die Prämie heraus.

{"**In Ihrem Report:** " + str(sh_count) + " Assignments erkannt (Call + Put), " + fmt_de(sh_eur) + " EUR Stillhalterprämien von Topf 1 nach Topf 2 verschoben." if sh_count > 0 else "**In Ihrem Report:** Keine Assignments erkannt."}

---

### Dividenden & Payment in Lieu (PIL)

- **Dividenden** (DIV): Laufende Erträge in Topf 2
- **Payment in Lieu** (PIL): Ersatzzahlung wenn Aktien verliehen sind  - wird steuerlich wie eine Dividende behandelt und mit diesen zusammen verrechnet

---

### Zinsen & Stückzinsen

- **INTR/CINT**: Zins- und Couponerträge aus Anleihen → Topf 2
- **INTP** (Stückzinsen): Beim Kauf einer Anleihe gezahlte aufgelaufene Zinsen sind **negative Einnahmen** (BMF Rn. 51). Sie reduzieren den Zinsertrag und können diesen insgesamt negativ werden lassen.

---

### Quellensteuer (Zeile 41)

Ausländische Quellensteuern auf Dividenden und Zinsen (z.B. 15% US-Quellensteuer) werden in Zeile 41 als **anrechenbare ausländische Steuern** gemeldet. Das Finanzamt rechnet diese nach der Verlustverrechnung auf die deutsche Abgeltungsteuer an.

---

### Zeilen-Zuordnung Anlage KAP

Da Interactive Brokers ein **ausländischer Broker ohne inländischen Steuerabzug** ist, werden die Einkünfte in der Sektion „Kapitalerträge, die **nicht** dem inländischen Steuerabzug unterlegen haben" (Zeilen 18–23) eingetragen:

| Zeile | Bedeutung | Berechnung |
|-------|-----------|------------|
| **19** | Ausländische Kapitalerträge (Netto) | Topf 1 + Topf 2 (Summe aller Erträge und Verluste) |
| **20** | Davon: Aktiengewinne | Brutto-Aktiengewinne (ohne Verluste) |
| **22** | Verluste ohne Aktien | Verluste aus Optionen, Futures, Anleihen etc. (positiver Betrag) |
| **23** | Aktienverluste | Verluste aus Aktienveräußerungen (positiver Betrag) |
| **41** | Anrechenbare Quellensteuer | Summe aller ausländischen Quellensteuern |

---

### Währungsumrechnung

{"**Ihr Konto hat EUR als Basiswährung.** Alle Beträge in der IBKR-Abrechnung sind bereits in EUR umgerechnet. Bei USD-Trades nutzt IBKR den Tageskurs (`fxRateToBase`), der direkt in EUR umrechnet  - kein zusätzlicher FX-Lookup erforderlich." if base_curr == "EUR" else "**Ihr Konto hat USD als Basiswährung.** Beträge werden in zwei Schritten umgerechnet: (1) Trade-Währung → USD über `fxRateToBase`, (2) USD → EUR über den Tageskurs des vorherigen Geschäftstags. Die täglichen USD/EUR-Kurse werden aus den IBKR-Daten extrahiert."}

---

### Rechtsgrundlagen

- **§20 EStG**  - Einkünfte aus Kapitalvermögen
- **BMF-Schreiben vom 14.05.2025**  - „Einzelfragen zur Abgeltungsteuer" (IV C 1 - S 2252/00075/016/070)
- **Jahressteuergesetz 2024**  - Abschaffung des €20.000-Caps für Termingeschäfteverluste (§20 Abs. 6 Satz 5 EStG), rückwirkend für alle offenen Fälle
- **Anlage KAP**  - Zeilen 9/14 (Termingeschäfte) existieren nur in der Sektion mit inländischem Steuerabzug und sind für IBKR nicht relevant
""")

with st.expander("Berechnungsdetails - So werden die XML-Daten verarbeitet"):
    st.markdown(f"""
### Schritt 1: XML-Extraktion

Die IBKR Flex Query XML wird in einzelne CSV-Dateien zerlegt. Jede XML-Sektion enthält spezifische Daten:

| XML-Sektion | Inhalt | Filter |
|---|---|---|
| `<Trades>` | Alle Trades. Felder: `assetCategory`, `fifoPnlRealized`, `fxRateToBase`, `reportDate`, `buySell`, `transactionType` | `EXECUTION` → trades.csv, `CLOSED_LOT` → closed_lots.csv (für Tageskurs-Korrektur) |
| `<StmtFunds>` | Dividenden, Zinsen, Steuern, Gebühren. Felder: `activityCode`, `amount`, `fxRateToBase`, `reportDate`, `transactionID` | Duplikate per `transactionID` entfernt |
| `<FIFOPerformanceSummaryInBase>` | Aggregierter PnL pro Instrument. Felder: `assetCategory`, `isin`, `totalRealizedPnl` | Fallback für fehlende Trades (z.B. T-Bill Maturity) |
| `<FxTransactions>` | FX-Gewinne/-Verluste. Felder: `fxCurrency`, `realizedPL`, `reportDate` | Nur `levelOfDetail=TRANSACTION` |
| `<AccountInformation>` | Basiswährung (`currency`), Kontotyp | Einzelner Eintrag |
| `<FlexStatement>` | Berichtszeitraum → Steuerjahr aus `toDate` | Automatisch erkannt |

**Multi-XML (Vorjahre):** Trades aus allen XMLs werden in eine gemeinsame `trades.csv` zusammengeführt (für Stillhalter-Matching über Jahresgrenzen). FX-Transaktionen werden chronologisch gemergt mit Deduplizierung per `transactionID`.

---

### Schritt 2: Deduplizierung

IBKR liefert in einigen Sektionen Duplikate:

| Quelle | Duplikat-Ursache | Deduplizierungs-Schlüssel |
|---|---|---|
| **Trades** | Erweiterte Flex Queries enthalten ORDER + EXECUTION für denselben Trade | `tradeID` (wenn vorhanden) oder `(dateTime, isin, buySell, quantity, closePrice, fifoPnlRealized)` |
| **StmtFunds** | IBKR bucht EUR-Transaktionen doppelt (Original-Währung + BaseCurrency-Ansicht) | `transactionID`. Erster Eintrag hat korrekten `fxRateToBase`, Duplikat hat `fxRateToBase=1` |

---

### Schritt 3: Kapitalgewinne berechnen

Für jeden Trade im Steuerjahr (`reportDate.year == Steuerjahr`):

```
IBKR-Methode:     PnL (EUR) = fifoPnlRealized × fxRateToBase
Tageskurs-Methode: PnL (EUR) = Erlös × FX_Verkaufstag − AK × FX_Kauftag
```

**IBKR-Methode (Standard):** Rechnet den Netto-PnL komplett zum Schlusskurs um.

**Tageskurs-Methode (§20 Abs. 4 S. 1 EStG, optional):** *"Bei nicht in Euro getätigten Geschäften sind die Einnahmen im Zeitpunkt der Veräußerung und die Anschaffungskosten im Zeitpunkt der Anschaffung in Euro umzurechnen."* Verwendet CLOSED_LOT Daten aus Extended Flex Queries. Futures werden ausgeschlossen (Kostenbasis = Notional, kein realer Cashflow). Korrektur: `|AK| × (FX_Schlusskurs - FX_Kaufkurs)` pro Lot. IBKR vergibt pro Tag zwei `fxRateToBase`-Kurse: einen Intraday-Kurs (ExchTrades) und einen Settlement-Kurs (BookTrades, 16:20). Für den Kaufkurs wird der ExchTrade-Kurs bevorzugt; an reinen Verfall-/Andienungstagen der BookTrade-Kurs als Fallback.

| Feld | Bedeutung |
|---|---|
| `fifoPnlRealized` | IBKR's FIFO-basierter realisierter Gewinn/Verlust in **Trade-Währung** |
| `fxRateToBase` | Umrechnungskurs Trade-Währung → Basiswährung (EUR) am **Schlusstag** |
| `reportDate` | Buchungsdatum (bestimmt das Steuerjahr, Zuflussprinzip) |
| `assetCategory` | Topf-Zuordnung: `STK` → Topf 1 oder KAP-INV, alles andere → Topf 2 |
| `subCategory` | ETF-Erkennung: `ETF` → InvStG-Prüfung, `COMMON` → Einzelaktie |

**Topf-Zuordnung:**

| `assetCategory` | `subCategory` | Steuerliche Einordnung | Topf |
|---|---|---|---|
| `STK` | `COMMON` / `REIT` / `ADR` | Aktienveräußerung (§20 Abs. 2 Nr. 1) | **Topf 1** |
| `STK` | `ETF` (InvStG-Fonds) | Investmentfonds (InvStG §2) | **KAP-INV** (optional) |
| `STK` | `ETF` (no\_invstg, z.B. IBIT) | Wie Einzelaktie | **Topf 1** |
| `OPT` | | Termingeschäft, Option (§20 Abs. 2 Nr. 3) | Topf 2 |
| `FUT` | | Termingeschäft, Future (§20 Abs. 2 Nr. 3) | Topf 2 |
| `FOP` / `FSFOP` | | Termingeschäft, Future-Option (§20 Abs. 2 Nr. 3) | Topf 2 |
| `BILL` | | Kapitalforderung, T-Bill (§20 Abs. 2 Nr. 7) | Topf 2 |
| `BOND` | | Kapitalforderung, Anleihe (§20 Abs. 2 Nr. 7) | Topf 2 |

**InvStG-Klassifizierung (optional):** ETFs mit `subCategory="ETF"` werden gegen eine Lookup-Tabelle (139 US-ETFs) geprüft. Aktienfonds (≥51% Aktienquote) erhalten 30% Teilfreistellung, sonstige Fonds 0%. Crypto/Commodity-ETPs (IBIT, GLD etc.) bleiben in Topf 1. Optionen auf ETFs bleiben in Topf 2.

**Jahresfilter:** Es wird `reportDate` verwendet, nicht `dateTime`. Grund: Trades am Jahresende (z.B. `dateTime=2024-12-29`, Settlement `reportDate=2025-01-02`) gehören steuerlich zum Settlement-Jahr. Dies entspricht dem Zuflussprinzip (§11 EStG).

---

### Schritt 4: Stillhalterprämien separieren (BMF Rn. 26, 33)

Bei Optionsassignments bündelt IBKR die Prämie in den Aktien-PnL. Das BMF verlangt eine Trennung:

**Erkennung eines Assignments:**
- `assetCategory` ∈ (OPT, FOP, FSFOP)
- `transactionType` = `BookTrade` (keine Börsentransaktion, sondern Ausbuchung)
- `buySell` = `BUY` (Short-Position wird geschlossen)
- `putCall` ∈ (C, P), sowohl Calls als auch Puts
- `fifoPnlRealized` ≈ 0 (IBKR zeigt keinen PnL auf der Option)

**Original-Verkauf finden:**
- Alle `ExchTrade SELL` mit identischem `strike`, `expiry`, `putCall`
- Können mehrere Teilfüllungen sein → gewichteter Durchschnitt

**Prämien-Berechnung:**
```
Prämie (Trade-Währung) = tradePrice × multiplier × quantity
Prämie (EUR) = Prämie × fxRateToBase (gewichtet über Teilfüllungen)
```

**Topf-Umbuchung:**
- `stocks_gain -= Prämie` (aus Topf 1 entfernen)
- `options_gain += Prämie` (in Topf 2 als §20 Abs. 1 Nr. 11)

**Cross-Year:** Wenn die Option in einem Vorjahr verkauft wurde (z.B. 2024) und im Steuerjahr (2025) assigned wird, gehört die Prämie ins Vorjahr (Zuflussprinzip). Vorjahres-XMLs müssen hochgeladen werden, damit der Original-SELL gefunden wird.

**Cross-Year Put-Korrektur:** Wenn Aktien aus Put-Assignments früherer Jahre im Steuerjahr verkauft werden, wird IBKR's PnL korrigiert. Die Prämie war bereits im Assignment-Jahr versteuert und darf die Anschaffungskosten nicht mindern. FIFO-Lot-Matching per Symbol.

---

### Schritt 5: Dividenden, Zinsen & Quellensteuer

Aus `statement_of_funds.csv` werden Cash-Positionen nach `activityCode` zugeordnet:

| `activityCode` | Bedeutung | Zuordnung |
|---|---|---|
| `DIV` | Dividenden | Topf 2 (§20 Abs. 1 Nr. 1) |
| `PIL` | Payment in Lieu (Ersatzzahlung bei Wertpapierleihe) | Wie Dividende, Topf 2 |
| `INTR` | Anleihekupon / Zinserträge | Topf 2 (§20 Abs. 1 Nr. 7) |
| `CINT` | Credit Interest (Guthabenzinsen) | Topf 2 |
| `INTP` | Stückzinsen (beim Kauf gezahlt) | Negative Einnahmen, Topf 2 (BMF Rn. 51) |
| `DINT` | Debit Interest (Sollzinsen, Leihgebühren, SYEP) | Negativ, Topf 2 |
| `FRTAX` / `WHT` | Quellensteuer (Withholding Tax) | Zeile 41 (anrechenbar) |

**Währungsumrechnung (EUR-Basis):** `amount` ist bereits in EUR (BaseCurrency-Ansicht). Keine weitere Umrechnung nötig.

**Jahresfilter:** `reportDate.year == Steuerjahr`. Steuer-Rückforderungen (Tax Reclaims) aus Vorjahren, die im Steuerjahr gebucht werden, sind korrekt dem Buchungsjahr zugeordnet.

---

### Schritt 6: Währungsumrechnung

{"**Ihr Konto: EUR-Basis.** Alle Beträge in `statement_of_funds` und `fifoPnlRealized × fxRateToBase` sind direkt in EUR. Es wird kein separater Tageskurs-Lookup benötigt." if base_curr == "EUR" else "**Ihr Konto: USD-Basis.** Zweistufige Umrechnung: (1) `fifoPnlRealized × fxRateToBase` → USD, (2) USD → EUR über täglichen Wechselkurs. Kurse werden primär aus EUR-Einträgen in Trades/Funds extrahiert; Lücken werden automatisch mit **EZB-Referenzkursen** gefüllt."}

| Szenario | Formel |
|---|---|
| **EUR-Base, EUR-Trade** | `PnL_EUR = fifoPnlRealized × fxRateToBase` (fxRate ≈ 1.0) |
| **EUR-Base, USD-Trade** | `PnL_EUR = fifoPnlRealized × fxRateToBase` (fxRate ≈ 0.86–0.92) |
| **USD-Base, USD-Trade** | `PnL_EUR = fifoPnlRealized × fxRateToBase × daily_usd_eur_rate` |
| **USD-Base, EUR-Trade** | `PnL_EUR = amount_eur` (direkt, da Trade in EUR) |

**Plausibilitätsprüfung:** USD→EUR-Kurse außerhalb [0.70, 1.30] werden verworfen. `fxRateToBase=1.0` auf EUR-Währungseinträgen wird als Duplikat übersprungen.

---

### Schritt 7: FX-Gewinne/-Verluste

Fremdwährungsgewinne/-verluste entstehen durch Kursänderungen auf verzinslichen Fremdwährungskonten (BMF Rn. 131, §20 Abs. 2 S. 1 Nr. 7 EStG).

**Datenquellen (Priorität):**

| Priorität | Quelle | Genauigkeit | Wann verfügbar |
|---|---|---|---|
| 1. | **XML `<FxTransactions>`** | Exakt (IBKR-internes FIFO, `realizedPL` pro Transaktion) | Wenn in Flex Query aktiviert |
| 2. | **IBKR Standard-Bericht (CSV)** | Exakt (gleiche Daten wie #1, aggregiert) | Manuell erstellt |
| 3. | **FIFO-Approximation** | Ungenau (~84% der Tageskurse unbrauchbar) | Immer (aus StmtFunds) |

FX-Gewinne/-Verluste fließen in **Topf 2**.

---

### Schritt 8: Anlage KAP + KAP-INV Berechnung

```
Topf 1 = Aktiengewinne + Aktienverluste (nach Stillhalter-Separation)
         (ohne InvStG-ETFs, wenn aktiviert)
Topf 2 = Dividenden + Zinsen + Optionsgewinne + Optionsverluste
         (inkl. Stillhalterprämien + FX-Gewinne/-Verluste)

Zeile 19 = Topf 1 + Topf 2 (Nettobetrag)
Zeile 20 = Aktiengewinne (brutto, ohne Verluste)
Zeile 22 = |Verluste ohne Aktien| (positiver Betrag)
Zeile 23 = |Aktienverluste| (positiver Betrag)
Zeile 41 = |Quellensteuer| (anrechenbar, positiver Betrag)
```

**Anlage KAP-INV (wenn InvStG aktiviert):**

```
ETF-Gewinne/-Verluste und ETF-Dividenden werden auf KAP-INV gemeldet.
Teilfreistellung wird pro ISIN angewendet:
  Aktienfonds (≥51% Aktienquote): 30% steuerfrei
  Sonstiger Fonds:                 0% steuerfrei

KAP-INV Netto = (ETF-G/V × (1 − TFS)) + (ETF-Div × (1 − TFS))
ETF-Quellensteuer wird separat auf KAP-INV angerechnet.
```

**Tageskurs-Korrektur (wenn aktiviert):**

```
Korrektur = Σ |Anschaffungskosten| × (FX_Verkauf − FX_Kauf) pro CLOSED_LOT
Futures ausgeschlossen (Kostenbasis = Notional, kein realer Cashflow).
Wird auf Topf 1, Topf 2 und KAP-INV aufgeteilt.
```
""")

# ── Export ───────────────────────────────────────────────────────────────────

section_title("Export")

# Build optional export sections
fx_export = ""
if fx_results:
    fx_export = "\nFREMDWÄHRUNGS-GEWINNE/VERLUSTE (FIFO)\n"
    for curr, data in sorted(fx_results.items()):
        fx_export += f"  {curr}: Gewinn {fmt_de(data['gain']):>10}  Verlust {fmt_de(data['loss']):>10}  Netto {fmt_de(data['net']):>10} EUR\n"
    fx_net = fx_total_gain + fx_total_loss
    fx_export += f"  ─────────────────────────────────────────────────\n"
    fx_export += f"  FX Gesamt Gewinn:      {fmt_de(fx_total_gain):>14} EUR\n"
    fx_export += f"  FX Gesamt Verlust:     {fmt_de(fx_total_loss):>14} EUR\n"
    fx_export += f"  FX Netto:              {fmt_de(fx_net):>14} EUR\n"
    fx_export += "  (In Topf 2 enthalten, BMF Rn. 131)\n"

sh_export = ""
if sh_count > 0:
    sh_export = f"\nSTILLHALTERPRÄMIEN (BMF Rn. 25-35)\n"
    sh_export += f"  {sh_count} Assignment(s) erkannt\n"
    sh_export += f"  Prämien umgebucht:     {fmt_de(sh_eur):>14} EUR\n"
    sh_export += f"  (Von Topf 1 nach Topf 2 verschoben)\n"

inv_export = ""
if has_etf_data and invstg_aktiv:
    inv_export = f"\nANLAGE KAP-INV: INVESTMENTFONDS (InvStG)\n"
    inv_export += f"  ETF-Gewinne (roh):     {fmt_de(etf_gain_raw):>14} EUR\n"
    inv_export += f"  ETF-Verluste (roh):    {fmt_de(etf_loss_raw):>14} EUR\n"
    inv_export += f"  ETF-Gewinne (stpfl.):  {fmt_de(etf_gain_taxable):>14} EUR\n"
    inv_export += f"  ETF-Verluste (stpfl.): {fmt_de(etf_loss_taxable):>14} EUR\n"
    inv_export += f"  ETF-Dividenden (roh):  {fmt_de(etf_div_raw):>14} EUR\n"
    inv_export += f"  ETF-Dividenden (stpfl.):{fmt_de(etf_div_taxable):>13} EUR\n"
    inv_export += f"  ETF-Quellensteuer:     {fmt_de(etf_wht):>14} EUR\n"
    inv_export += f"  ─────────────────────────────────────────────────\n"
    inv_export += f"  KAP-INV Netto (stpfl.):{fmt_de(etf_net_taxable):>14} EUR\n"
    for isin, info in sorted(etf_by_isin.items(), key=lambda x: x[1].get('ticker', '')):
        gv_tax = info.get('gain_taxable', 0) + info.get('loss_taxable', 0)
        inv_export += f"    {info.get('ticker', isin):8s} TFS {info.get('tfs_rate', 0)*100:.0f}%  G/V stpfl. {fmt_de(gv_tax):>10} EUR\n"

topf2_detail_export = ""
if topf2_cats:
    topf2_detail_export = "\nAUFSCHLÜSSELUNG TOPF 2\n"
    div_eur = d.get('dividends_eur', 0)
    int_eur = d.get('interest_eur', 0)
    topf2_detail_export += f"  {'Dividenden':24s} G {fmt_de(max(div_eur, 0)):>10} V {fmt_de(min(div_eur, 0)):>10} N {fmt_de(div_eur):>10} EUR\n"
    topf2_detail_export += f"  {'Zinsen':24s} G {fmt_de(max(int_eur, 0)):>10} V {fmt_de(min(int_eur, 0)):>10} N {fmt_de(int_eur):>10} EUR\n"
    for cat, vals in sorted(topf2_cats.items()):
        net = vals['gain'] + vals['loss']
        topf2_detail_export += f"  {cat:24s} G {fmt_de(vals['gain']):>10} V {fmt_de(vals['loss']):>10} N {fmt_de(net):>10} EUR\n"

multi_acct_export = ""
if n_accounts > 1:
    multi_acct_export = f"Konten: {n_accounts} (separat berechnet, Ergebnisse addiert)\n"

report_text = f"""ANLAGE KAP {steuerjahr} - Steuerbericht
Erstellt: {_dt.now().strftime('%d.%m.%Y %H:%M')}
Basiswährung: {d.get('base_currency', 'USD')}
{multi_acct_export}

═══════════════════════════════════════════════════
TOPF 1: AKTIEN (ohne ETF-Fonds)
  Aktiengewinne:         {fmt_de(d.get('stocks_gain_eur', 0)):>14} EUR
  Aktienverluste:        {fmt_de(d.get('stocks_loss_eur', 0)):>14} EUR
  ─────────────────────────────────────────────────
  Saldo Aktien:          {fmt_de(topf_1):>14} EUR

TOPF 2: SONSTIGES (inkl. Termingeschäfte)
  Dividenden:            {fmt_de(d.get('dividends_eur', 0)):>14} EUR
  Zinsen (netto):        {fmt_de(d.get('interest_eur', 0)):>14} EUR
  Sonstige Gewinne:     {fmt_de(d.get('options_gain_eur', 0)):>14} EUR
  Sonstige Verluste:    {fmt_de(d.get('options_loss_eur', 0)):>14} EUR
  ─────────────────────────────────────────────────
  Saldo Sonstiges:       {fmt_de(topf_2):>14} EUR
{topf2_detail_export}{fx_export}{sh_export}{inv_export}
═══════════════════════════════════════════════════
ANLAGE KAP EINTRAGUNGEN
  Zeile 19 (Netto):      {fmt_de(zeile_19):>14} EUR
  Zeile 20 (Aktiengewinne): {fmt_de(zeile_20):>11} EUR
  Zeile 22 (Verluste o. Aktien): {fmt_de(zeile_22):>8} EUR
  Zeile 23 (Aktienverluste): {fmt_de(zeile_23):>11} EUR
  Zeile 41 (Quellensteuer): {fmt_de(quellensteuer):>11} EUR
{"" if not (has_etf_data and invstg_aktiv) else chr(10) + "ANLAGE KAP-INV EINTRAGUNGEN" + chr(10) + f"  KAP-INV Erträge (nach TFS): {fmt_de(etf_net_taxable):>8} EUR" + chr(10) + f"  KAP-INV Quellensteuer: {fmt_de(etf_wht):>13} EUR" + chr(10)}{"" if not has_so_data else chr(10) + "ANLAGE SO (§23 EStG) — PRIVATE VERÄUSSERUNGSGESCHÄFTE" + chr(10) + f"  Physische Gold-ETCs (BFH VIII R 4/15)" + chr(10) + f"  Steuerpflichtig (≤ 1J): {fmt_de(so_taxable_for_row):>12} EUR  → Anlage SO" + chr(10) + f"  Steuerfrei (> 1J):      {fmt_de(so_free_for_row):>12} EUR" + chr(10)}═══════════════════════════════════════════════════
"""

st.download_button(
    label="Textreport herunterladen",
    data=report_text,
    file_name=f"steuerbericht_{steuerjahr}.txt",
    mime="text/plain",
    use_container_width=True
)

# ── Rechtliche Hinweise ──────────────────────────────────────────────────────

section_title("Rechtliche Hinweise")

st.markdown("""
<div style="background: rgba(255,255,255,0.02); border: 1px solid rgba(255,255,255,0.06); border-radius: 12px; padding: 1.25rem 1.25rem; font-size: 0.78rem; color: #64748b; line-height: 1.7;">

<strong style="color: #94a3b8;">Haftungsausschluss</strong><br>
Dieses Tool dient ausschließlich zur Unterstützung bei der Erstellung der Einkommensteuererklärung. Die berechneten Werte sind unverbindlich und ohne Gewähr für Richtigkeit, Vollständigkeit oder Aktualität. Der Nutzer ist für die Prüfung aller Angaben in seiner Steuererklärung selbst verantwortlich. Die Nutzung erfolgt auf eigenes Risiko.

<br><br>
<strong style="color: #94a3b8;">Keine Steuerberatung</strong><br>
Dieses Tool stellt keine Steuerberatung im Sinne des Steuerberatungsgesetzes (StBerG) dar und ersetzt nicht die Beratung durch einen Steuerberater, Wirtschaftsprüfer oder eine andere zur Steuerberatung befugte Person. Bei Unsicherheiten oder komplexen Sachverhalten konsultieren Sie bitte einen steuerlichen Berater.

<br><br>
<strong style="color: #94a3b8;">Keine Haftung</strong><br>
Die Entwickler und Mitwirkenden dieses Projekts haften nicht für Schäden, die durch die Nutzung oder Nichtnutzung der berechneten Informationen entstehen, einschließlich, aber nicht beschränkt auf finanzielle Verluste, Steuernachzahlungen, Bußgelder oder Zinsen. Dies gilt sowohl für direkte als auch für indirekte Schäden, unabhängig davon, ob diese vorhersehbar waren.

<br><br>
<strong style="color: #94a3b8;">Datenschutz und Datenverarbeitung</strong><br>
Sämtliche Berechnungen werden ausschließlich lokal im Browser des Nutzers ausgeführt (clientseitige Verarbeitung mittels WebAssembly). Es werden zu keinem Zeitpunkt personenbezogene Daten, Finanzdaten oder hochgeladene Dateien an einen Server übertragen, gespeichert oder an Dritte weitergegeben. Es findet kein Tracking, keine Analyse und keine Protokollierung statt. Die Anwendung erfüllt die Anforderungen der DSGVO, da keine Datenverarbeitung durch den Anbieter erfolgt.

<br><br>
<strong style="color: #94a3b8;">Rechtsstand und Aktualität</strong><br>
Die steuerlichen Berechnungen basieren auf dem Rechtsstand des Steuerjahres 2025, insbesondere auf §20 EStG, dem BMF-Schreiben vom 14.05.2025 (Einzelfragen zur Abgeltungsteuer) sowie dem Jahressteuergesetz 2024. Änderungen der Rechtslage, der Verwaltungsauffassung oder der Rechtsprechung nach Veröffentlichung dieses Tools werden nicht automatisch berücksichtigt.

<br><br>
<strong style="color: #94a3b8;">Open Source</strong><br>
Dieses Projekt ist unter der MIT-Lizenz veröffentlicht. Der Quellcode ist frei einsehbar und prüfbar unter
<a href="https://github.com/KonvexInvestment/ibkr-steuer" target="_blank" style="color: #60a5fa;">github.com/KonvexInvestment/ibkr-steuer</a>.
Jeder kann den Code einsehen, prüfen und zur Verbesserung beitragen.

</div>
""", unsafe_allow_html=True)
